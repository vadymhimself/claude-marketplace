"""
Phase 6 — Build dark-mode xlsx workbook for Ubiquify audit.

Sheets:
 1. Executive Summary (three-tier WINS / OKAY / CRITICAL)
 2. Retro Evidence
 3. Competitive Deep-Dive
 4. Chat Excerpts (status = skipped — leads.chats dry)
 5. Win/Loss CL Table
 6. Auto-Bidding Aggregates
 7. Recommendations Detail
"""
import json, os, hashlib, io
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.drawing.image import Image as XLImage

# Image download cache for profile avatars — avoid re-downloading on repeated
# build_workbook runs. Images are fetched from Upwork's public CDN and resized
# to fit inside ~80x80 px cells via Pillow before openpyxl embeds them.
_IMG_CACHE_DIR = Path("/tmp/profile_imgs")
_IMG_CACHE_DIR.mkdir(parents=True, exist_ok=True)

def _download_and_size_image(url, size=(80, 80)):
    """Download an image URL (Upwork CDN) → resize → cache locally.
    Returns an openpyxl.drawing.image.Image ready to add_image(), or None on
    any failure (404, timeout, unsupported format). Call sites must handle
    None by rendering plain text instead."""
    if not url or not isinstance(url, str) or not url.startswith("http"):
        return None
    url_hash = hashlib.md5(url.encode()).hexdigest()[:16]
    cached = _IMG_CACHE_DIR / f"{url_hash}.png"
    if not cached.exists():
        try:
            import requests, urllib3
            urllib3.disable_warnings()
            r = requests.get(url, timeout=8, verify=False,
                             headers={"User-Agent": "Mozilla/5.0"})
            if r.status_code != 200:
                return None
            raw = r.content
        except Exception:
            return None
        try:
            from PIL import Image as PILImage
            im = PILImage.open(io.BytesIO(raw))
            if im.mode not in ("RGB", "RGBA"):
                im = im.convert("RGBA")
            im.thumbnail(size, PILImage.LANCZOS)
            im.save(cached, "PNG")
        except Exception:
            try:
                cached.write_bytes(raw)  # fallback: raw bytes (openpyxl may still load)
            except Exception:
                return None
    try:
        img = XLImage(str(cached))
        # Rough size hint — openpyxl honours these as pixel dimensions.
        img.width = size[0]
        img.height = size[1]
        return img
    except Exception:
        return None

# ---- palette ----
BG = "0F1419"
BG_ALT = "161B22"
FG = "E6EDF3"
FG_DIM = "8B949E"
WIN_BG = "1F6F3F"
WIN_FG = "9EF0B2"
OKAY_BG = "7A5A1A"
OKAY_FG = "F3D582"
CRIT_BG = "7A2021"
CRIT_FG = "F1949A"
HEADER_BG = "21262D"
LINK_FG = "58A6FF"

FONT = "Calibri"
MONO = "Consolas"

def fg(color, size=10):
    return Font(name=FONT, color=color, size=size)

def fg_bold(color, size=10):
    return Font(name=FONT, color=color, size=size, bold=True)

def fg_mono(color, size=9):
    return Font(name=MONO, color=color, size=size)

def fill(color):
    return PatternFill("solid", start_color=color, end_color=color)

thin = Side(style="thin", color=FG_DIM)
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

# ---- load phase data ----
def ld(name):
    with open(f"/sessions/dazzling-nifty-fermat/audit_work/v2_ubiquify/{name}") as f:
        return json.load(f)

phase1 = ld("phase1_retro_v2.json")
phase2a = ld("phase2a_cohort_v2.json")
phase2b = ld("phase2b_peer_knn_v2.json")          # enriched v2 — KNN cohort + named agencies + freelancers + CL templates
phase2c = ld("phase2c_match_reasoning.json")      # per-competitor-win → UB analogue pair + reasoning paragraphs
phase3 = ld("phase3_chats.json")
phase4 = ld("phase4_winloss.json")
phase5 = ld("phase5_aggregates.json")

UB_HL_BG = "203040"  # Ubiquify highlight row background
UB_HL_FG = "79C0FF"  # Ubiquify highlight text


wb = Workbook()

def darken_sheet(ws):
    ws.sheet_view.showGridLines = False
    # Dark tab color
    ws.sheet_properties.tabColor = "0F1419"

def fill_row_bg(ws, row, start_col, end_col, color=BG):
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=c)
        if cell.fill.fill_type != "solid" or cell.fill.start_color.rgb != color:
            cell.fill = fill(color)

def fill_sheet_bg(ws, max_row=200, max_col=30):
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if not cell.fill.fill_type or cell.fill.fill_type == "none":
                cell.fill = fill(BG)
                cell.font = fg(FG)

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# =========================================================================
# Sheet 1 — Executive Summary
# =========================================================================
s1 = wb.active
s1.title = "Executive Summary"
darken_sheet(s1)

# Title
s1["A1"] = "GigRadar Audit — Ubiquify Digital"
s1["A1"].font = Font(name=FONT, size=22, bold=True, color=FG)
s1["A1"].fill = fill(BG)
s1.row_dimensions[1].height = 32
s1["A2"] = f"Focus window: 2026-03-23 to 2026-04-22 (30 days)   •   Team: daniyal@ubiquifydigital.com"
s1["A2"].font = fg(FG_DIM)
s1["A3"] = "Methodology v2 — reply rate + $/reply as north-star metrics. Hire rate diagnostic only (off-Upwork closes undercount wins)."
s1["A3"].font = fg(FG_DIM)

row = 5

# ---------- Headline metrics box ----------
fw = phase1["focus_window"]["auto_bidder"]
pw = phase1["prior_window"]["auto_bidder"]
inf = phase2a["inferred_cohort"]
brd = phase2a["broad_cohort"]
knn = phase2b["knn_cohort_summary"]

s1.cell(row=row, column=1, value="HEADLINE METRICS (focus window, outbound auto-bidder)").font = fg_bold(FG, 14)
s1.cell(row=row, column=1).fill = fill(HEADER_BG)
fill_row_bg(s1, row, 1, 7, HEADER_BG)
row += 1

headline_headers = ["Metric", "Value", "Prior Window", "Cohort Median (N=157 peers)", "Percentile (inferred peers)", "Percentile (all 210 teams ≥100 sent)", f"Percentile (KNN territorial N={knn['qualified_count']})"]
for i, h in enumerate(headline_headers, 1):
    c = s1.cell(row=row, column=i, value=h)
    c.font = fg_bold(FG_DIM)
    c.fill = fill(HEADER_BG)
row += 1

hl_rows = [
    ("Reply rate", f"{fw['reply_rate']*100:.2f}%", f"{pw['reply_rate']*100:.2f}%",
        f"{inf['reply_rate_stats']['median']*100:.2f}%",
        f"P{inf['subject_position']['reply_rate_percentile']:.0f}",
        f"P{brd['subject_position']['reply_rate_percentile']:.0f}",
        f"P{knn['subject_position']['reply_rate_percentile']:.0f}"),
    ("$ / reply",  f"${fw['cost_per_reply_usd']:.2f}", f"${pw['cost_per_reply_usd']:.2f}",
        f"${inf['cost_per_reply_stats']['median']:.2f}",
        f"P{inf['subject_position']['cost_per_reply_percentile_lower_better']:.0f}",
        f"P{brd['subject_position']['cost_per_reply_percentile_lower_better']:.0f}",
        f"P{knn['subject_position']['cost_per_reply_percentile_lower_better']:.0f}"),
    ("Hire rate (diagnostic)", f"{fw['hire_rate']*100:.2f}%", f"{pw['hire_rate']*100:.2f}%",
        f"{inf['hire_rate_among_hiring_stats']['median']*100:.2f}% (among N=46 hiring peers)",
        f"P{inf['subject_position'].get('hire_rate_percentile_among_hiring', 0):.0f}",
        f"P{brd['subject_position'].get('hire_rate_percentile_among_hiring', 0):.0f}",
        f"P{knn['subject_position'].get('hire_rate_percentile_among_hiring', 0):.0f}"),
    ("Sent (outbound auto)", f"{fw['sent']}", f"{pw['sent']}", "—", "—", "—", "—"),
    ("Connects spend", f"${fw['connects_spend_usd']:.2f}", f"${pw['connects_spend_usd']:.2f}", "—", "—", "—", "—"),
]
for r in hl_rows:
    for i, v in enumerate(r, 1):
        c = s1.cell(row=row, column=i, value=v)
        c.font = fg(FG) if i != 1 else fg_bold(FG)
        c.fill = fill(BG_ALT)
        c.alignment = Alignment(horizontal="left", vertical="center")
    row += 1

row += 1

# ---------- Cohort split ----------
s1.cell(row=row, column=1, value="COHORT SPLIT (focus window)").font = fg_bold(FG, 14)
s1.cell(row=row, column=1).fill = fill(HEADER_BG)
fill_row_bg(s1, row, 1, 6, HEADER_BG)
row += 1
split_headers = ["Cohort", "Sent", "Replied", "Hired", "Reply Rate", "Hire Rate"]
for i, h in enumerate(split_headers, 1):
    c = s1.cell(row=row, column=i, value=h)
    c.font = fg_bold(FG_DIM)
    c.fill = fill(HEADER_BG)
row += 1

# all-time cohort data
ats = phase1["all_time"]
fws = phase1["focus_window"]

split_rows = [
    ("Auto-bidder (outbound, focus)", fws["auto_bidder"]["sent"], fws["auto_bidder"]["replied"], fws["auto_bidder"]["hired"],
     f"{fws['auto_bidder']['reply_rate']*100:.2f}%",
     f"{fws['auto_bidder']['hire_rate']*100:.2f}%"),
    ("Manual outbound (focus)", fws["manual"]["sent"], fws["manual"]["replied"], fws["manual"]["hired"],
     f"{fws['manual']['reply_rate']*100:.2f}%",
     f"{fws['manual']['hire_rate']*100:.2f}%"),
    ("All-time auto-bidder (cumulative)", ats["auto_bidder"]["sent"], ats["auto_bidder"]["replied"], ats["auto_bidder"]["hired"],
     f"{ats['auto_bidder']['reply_rate']*100:.2f}%",
     f"{ats['auto_bidder']['hire_rate']*100:.2f}%"),
    ("All-time manual-outbound", ats["manual"]["sent"], ats["manual"]["replied"], ats["manual"]["hired"],
     f"{ats['manual']['reply_rate']*100:.2f}%",
     f"{ats['manual']['hire_rate']*100:.2f}%"),
    ("All-time manual-INVITES (separate channel)", 29, 29, 3,
     "100.00%", "10.34%"),
]
for r in split_rows:
    for i, v in enumerate(r, 1):
        c = s1.cell(row=row, column=i, value=v)
        c.font = fg(FG) if i != 1 else fg_bold(FG)
        c.fill = fill(BG_ALT)
    row += 1

row += 1
s1.cell(row=row, column=1,
        value="Note: apples-to-apples, auto-outbound (8.58% RR all-time) and manual-outbound (8.57% RR all-time) reply rates are effectively IDENTICAL. Manual invites are a separate channel (100% RR by construction) and drive 12% of hires at 0.3% of volume."
).font = Font(name=FONT, size=9, italic=True, color=FG_DIM)
s1.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
s1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
s1.row_dimensions[row].height = 32
row += 2


# ---------- Tier 1 WINS ----------
s1.cell(row=row, column=1, value="TIER 1 — WINS (what's working; keep doing this)").font = fg_bold(WIN_FG, 14)
s1.cell(row=row, column=1).fill = fill(WIN_BG)
fill_row_bg(s1, row, 1, 6, WIN_BG)
row += 1

wins = [
    ("Reply-rate is ABOVE median peers",
     f"Focus 5.88% vs inferred-peer median 4.24% (P70). $/reply $51.77 vs peer median $71.82 — you're replying MORE than most peers and paying LESS per reply than most.",
     "Keep the account at current send cadence (~17/day); avoid expansion that dilutes targeting.",
     "Phase 2A Cohort Compare"),
    ("Scanner AM-05-C (newer algorithm) — 10.20% reply rate, $30.78/reply",
     "49 proposals, 5 replies in focus. Prior window same scanner was 7.87% — IMPROVING. Specific slice: AM-05-C × promptVersion=1.2.6.1 × ALG_LAZ algo. Best slice in team's entire portfolio.",
     "This is the prompt/algo combination that works — replicate onto AM-01/03/04 siblings. Preserve the AM-05-C query as-is.",
     "Phase 5 row 1 + Phase 4 Win/Loss"),
    ("n8n/zap/make scanner — 7.69% reply rate, $29.92/reply",
     "26 proposals, 2 replies. Cheapest $/reply on the whole account. Niche automation-agency positioning is working.",
     "Keep this scanner live. Study the exact boolean query and port the 'automation agency' angle to adjacent scanners.",
     "Phase 5 row 4"),
    ("Month-over-month improvement on reply rate",
     f"Focus 5.88% vs prior 5.35% (+0.53pp). Auto-bidder shipping real gains week-over-week.",
     "Continue current experimentation cadence (new algo + prompt version rollouts).",
     "Phase 1 Retro"),
    ("Invite channel is a hidden gold mine",
     "29 invite proposals in all-time → 100% reply rate → 3 hires (12% hire rate on invites vs 0.25% on outbound). This is not a scanner problem, it's a channel-discovery finding.",
     "Ensure invite notifications go to the freelancer fast (sub-hour). Consider dedicated invite-reply template.",
     "Phase 1 Retro invite split"),
]
for w in wins:
    s1.cell(row=row, column=1, value=w[0]).font = fg_bold(WIN_FG)
    s1.cell(row=row, column=2, value=w[1]).font = fg(FG)
    s1.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    s1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    s1.cell(row=row, column=5, value="Action: " + w[2]).font = fg(WIN_FG)
    s1.cell(row=row, column=5).alignment = Alignment(wrap_text=True, vertical="top")
    s1.cell(row=row, column=6, value=w[3]).font = fg(FG_DIM)
    fill_row_bg(s1, row, 1, 6, BG_ALT)
    s1.row_dimensions[row].height = 48
    row += 1

row += 1

# ---------- Tier 2 OKAY ----------
s1.cell(row=row, column=1, value="TIER 2 — OKAY (near-benchmark; concrete opportunities)").font = fg_bold(OKAY_FG, 14)
s1.cell(row=row, column=1).fill = fill(OKAY_BG)
fill_row_bg(s1, row, 1, 6, OKAY_BG)
row += 1

okays = [
    ("$/reply is above peer p25 but below median",
     f"$51.77 vs inferred-peer p25 $37.40 — there's ~28% room. Top-quartile peers (e.g. competitors RipeSeed, Vizio) spend ~$30-37/reply.",
     "Reduce connects bid on scanners with reply rate < 4% (AM-07-D 0%, V. python 4.65% but $60/reply, A4 Awais $128/reply). Cap connectsBid at 12 for scanners below team avg.",
     "Phase 5 + Phase 2B"),
    ("V. python scanner — 4.65% RR, $60/reply",
     "43 proposals, 2 replies in focus. Highest-volume scanner after AM-05-C but ~2x the cost. No hires in focus.",
     "A/B test V. python with the ALG_LAZ algo+1.2.6.2.mem prompt that's working on AM-05-C. Same scanner query, swap the CL generator.",
     "Phase 4 Win/Loss + Phase 5"),
    ("A4 Awais - Full Stack & AI — 2.63% RR, $128/reply",
     "38 proposals, 1 reply in focus. Most expensive $/reply in your portfolio for a high-volume scanner. Algo=None (default 'ㅤ⁤'), Prompt=1.2.6.1.",
     "Migrate A4 to the same algo+prompt combo used on AM-05-C. If RR doesn't move in 30 days, retire this scanner.",
     "Phase 4 + Phase 5"),
    ("Hire rate 0.59% — above peer median 0.53% but well below p75 1.3%",
     "Peers with more hires are converting replies→interviews better. This is the CL→interview funnel, not the pitch→reply funnel.",
     "Add a scripted first-response template optimized for 'quick call ask' — top peers (e.g. RipeSeed, WeSoftYou) include a Calendly link in their first chat message.",
     "Phase 2B Competitor Harvest"),
    ("2 algorithm versions running in parallel without A/B separation",
     "Default 'ㅤ⁤' algo + ALG_LAZ (only 3 samples). Prompt 1.2.6.1 (dominant) + 1.2.6.2.mem (3 samples). Not enough volume on newer configs to conclude.",
     "Route 50% of AM-05-C/AM-01-C/V.python/n8n-zap-make volume to ALG_LAZ+1.2.6.2.mem for 30 days. Collect proper A/B data.",
     "Phase 5"),
]
for w in okays:
    s1.cell(row=row, column=1, value=w[0]).font = fg_bold(OKAY_FG)
    s1.cell(row=row, column=2, value=w[1]).font = fg(FG)
    s1.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    s1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    s1.cell(row=row, column=5, value="Action: " + w[2]).font = fg(OKAY_FG)
    s1.cell(row=row, column=5).alignment = Alignment(wrap_text=True, vertical="top")
    s1.cell(row=row, column=6, value=w[3]).font = fg(FG_DIM)
    fill_row_bg(s1, row, 1, 6, BG_ALT)
    s1.row_dimensions[row].height = 54
    row += 1

row += 1

# ---------- Tier 3 CRITICAL ----------
s1.cell(row=row, column=1, value="TIER 3 — CRITICAL (must-fix; wasting spend)").font = fg_bold(CRIT_FG, 14)
s1.cell(row=row, column=1).fill = fill(CRIT_BG)
fill_row_bg(s1, row, 1, 6, CRIT_BG)
row += 1

crits = [
    ("D3 Awais - Full Stack & AI — 21 proposals, 0% reply rate",
     "Zero replies in focus on 21 sent. Previous window also 0. Wasted ~$63 on this scanner this month alone. The 'D3' template variant is not converting.",
     "PAUSE D3 Awais scanner immediately. Rewrite the boolean query or merge into A4/AM-05 before re-enabling.",
     "Phase 5 / Phase 4"),
    ("AM-07-D — 15 proposals, 0% reply rate (was 2.33% prior)",
     "Dropped from 2.33% → 0%. The -D sibling variant is regressing. Scanner config likely needs review.",
     "Pause AM-07-D. Compare its boolean query to AM-07-C and revert to the -C baseline.",
     "Phase 5 / Phase 4"),
    ("AM-05-A regressed to 0% (was 12% prior window)",
     "11 proposals, 0 replies. Previous window same scanner was 12% RR — a huge regression.",
     "Investigate: did the scanner query change? Did algo+prompt switch? Roll back AM-05-A to the config that had 12% RR.",
     "Phase 5"),
    ("AM-03-C regressed to 0% (was 8.33% prior)",
     "13 proposals, 0 replies. Another -C sibling dropping despite -05-C going up.",
     "Diff the AM-03-C vs AM-05-C boolean queries. AM-05-C is working — port its recent changes to AM-03-C.",
     "Phase 5"),
    ("Team serviceNames + industry fields empty on team document",
     "Ubiquify has no serviceNames[] or industry populated, which breaks GigRadar's built-in cohort benchmarking for the team. Dashboard benchmarks will silently skip.",
     "Populate teams.serviceNames = ['Web/Mobile/SW Dev','AI/ML'] and teams.industry so internal benchmark widgets work.",
     "Phase 1 / Phase 2A"),
    ("Leads chat-sync is not running for Ubiquify (759 proposals with chatId, 0 leads.chats docs)",
     "Cannot diagnose CL→interview→hire funnel because transcripts aren't synced. Missing signal for audit, account health, and churn prediction.",
     "Run the chat-sync backfill for team 679a215568faa05722aabb93. This is an internal GigRadar engineering ticket.",
     "Phase 3 (skipped)"),
]
for w in crits:
    s1.cell(row=row, column=1, value=w[0]).font = fg_bold(CRIT_FG)
    s1.cell(row=row, column=2, value=w[1]).font = fg(FG)
    s1.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    s1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    s1.cell(row=row, column=5, value="Action: " + w[2]).font = fg(CRIT_FG)
    s1.cell(row=row, column=5).alignment = Alignment(wrap_text=True, vertical="top")
    s1.cell(row=row, column=6, value=w[3]).font = fg(FG_DIM)
    fill_row_bg(s1, row, 1, 6, BG_ALT)
    s1.row_dimensions[row].height = 54
    row += 1

# =========================================================================
# Competitive Deep-Dive summary block (Sheet 1)
# Pulls priority tactics from phase2c judgments + one punchy line per
# competitor so the exec read is self-contained (no sheet-hopping required).
# =========================================================================
row += 1
s1.cell(row=row, column=1,
        value="COMPETITIVE DEEP-DIVE — PRIORITY TACTICS (cross-competitor)").font = fg_bold(UB_HL_FG, 14)
s1.cell(row=row, column=1).fill = fill(HEADER_BG)
fill_row_bg(s1, row, 1, 6, HEADER_BG)
row += 1
s1.cell(row=row, column=1,
        value="What winning competitors do on the SAME job type that Ubiquify does NOT. Full detail: Competitive Deep-Dive sheet · Full playbook: COMPETITIVE_PLAYBOOK.md (workspace root).").font = fg(FG_DIM)
s1.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
s1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
row += 1

# Tactic ranking (v0.4) — weight each tactic by the reply_rate of the competitor
# that authored it, then sum weights across all competitors mentioning a similar
# tactic. Higher-performing competitors' tactics naturally float to the top.
# Reply rate is preferred over raw count because a tactic from a 17% RR winner
# teaches us more than one from a 4% RR also-ran.
from collections import defaultdict
_tactic_weight = defaultdict(float)     # key -> accumulated reply-rate weight
_tactic_by_key = {}                     # key -> canonical body
_tactic_mentions = defaultdict(list)    # key -> list of (team_name, rr) tuples
_by_competitor = []                     # (team_name, rr, top_tactic) rows

# Index focus-window metrics by team_id for quick lookup
_comp_fm = {c.get("team_id"): (c.get("focus_metrics") or {}) for c in (phase2c.get("competitors") or [])}

for comp in (phase2c.get("competitors") or []):
    name = comp.get("team_name") or comp.get("team_id", "?")[:20]
    tid = comp.get("team_id")
    rr = (_comp_fm.get(tid) or {}).get("reply_rate") or 0.0
    tactics = comp.get("ai_tactics") or []
    if tactics:
        _by_competitor.append((name, rr, tactics[0]))
        for t in tactics:
            key = " ".join((t or "").split()[:8]).lower()
            _tactic_weight[key] += rr  # weight by reply rate, not count
            _tactic_by_key.setdefault(key, t)
            _tactic_mentions[key].append((name, rr))

# Top tactics ranked by a COMPOSITE weight of the authoring-competitor's
# focus-window performance. Competitors are SELECTED by neighborhood overlap
# (they must actually compete in the subject team's job space — see phase2b),
# but within that set we weight each competitor's advice by how well their
# strategy is actually working. Score =
#     reply_rate        (headline engagement)
#   + 2 × hire_rate     (conversion doubles since closes are the real win)
#   + min(hires/10, 0.5) (absolute-volume bonus — 20% hire rate off 1 hire
#                         is noisy, off 10 hires is a real signal)
# Tactics flatten 1-per-entry then sort by weight DESC; top 15 shown.
def _tactic_weight(rr, hr, hires):
    rr = rr or 0.0
    hr = hr or 0.0
    h = hires or 0
    return rr + 2.0 * hr + min(h / 10.0, 0.5)

all_tactics_flat = []  # (weight, rr, hr, hires, name, body)
for comp in (phase2c.get("competitors") or []):
    name = comp.get("team_name") or comp.get("team_id", "?")[:20]
    tid = comp.get("team_id")
    m = _comp_fm.get(tid) or {}
    rr = m.get("reply_rate") or 0.0
    hr = m.get("hire_rate") or 0.0
    hires = m.get("hired") or 0
    w = _tactic_weight(rr, hr, hires)
    for t in (comp.get("ai_tactics") or []):
        if t and isinstance(t, str):
            all_tactics_flat.append((w, rr, hr, hires, name, t))
top_tactics_flat = sorted(all_tactics_flat, key=lambda x: -x[0])[:15]

if top_tactics_flat:
    s1.cell(row=row, column=1, value="Ranked by composite score (RR + 2·HR + hire-count bonus)").font = fg_bold(FG, 11)
    s1.cell(row=row, column=2, value="Tactic").font = fg_bold(FG, 11)
    s1.cell(row=row, column=6, value="Source · RR · HR · hires · score").font = fg_bold(FG_DIM, 11)
    fill_row_bg(s1, row, 1, 6, BG_ALT)
    row += 1
    for w, rr, hr, hires, name, body in top_tactics_flat:
        rr_txt = f"{rr*100:.1f}%" if rr else "—"
        hr_txt = f"{hr*100:.1f}%" if hr else "0%"
        s1.cell(row=row, column=1, value=f"▸ {body[:60]}").font = fg_bold(UB_HL_FG)
        s1.cell(row=row, column=2, value=body).font = fg(FG)
        s1.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        s1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        src_txt = f"{name}\nRR {rr_txt} · HR {hr_txt} · {hires} hires · score {w:.2f}"
        s1.cell(row=row, column=6, value=src_txt).font = fg_bold(WIN_FG)
        s1.cell(row=row, column=6).alignment = Alignment(wrap_text=True, vertical="top")
        fill_row_bg(s1, row, 1, 6, BG_ALT)
        s1.row_dimensions[row].height = 70
        row += 1

# Per-competitor block — ordered by reply rate so the highest-RR competitor's
# top tactic sits at the top of this block.
if _by_competitor:
    row += 1
    s1.cell(row=row, column=1, value="Per-competitor top move (ranked by reply rate)").font = fg_bold(FG, 11)
    s1.cell(row=row, column=2, value="Strongest single tactic").font = fg_bold(FG, 11)
    fill_row_bg(s1, row, 1, 6, BG_ALT)
    row += 1
    _by_competitor_sorted = sorted(_by_competitor, key=lambda x: -x[1])
    for name, rr, tactic in _by_competitor_sorted[:10]:
        rr_txt = f"{rr*100:.1f}% RR" if rr else "—"
        s1.cell(row=row, column=1, value=f"{name}\n{rr_txt}").font = fg_bold(UB_HL_FG)
        s1.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        s1.cell(row=row, column=2, value=tactic).font = fg(FG)
        s1.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        s1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        fill_row_bg(s1, row, 1, 6, BG_ALT)
        s1.row_dimensions[row].height = 54
        row += 1

# fill remaining bg
fill_sheet_bg(s1, max_row=row + 5, max_col=7)

set_col_widths(s1, [32, 14, 14, 22, 24, 24, 24])


# =========================================================================
# Sheet 2 — Retro Evidence
# =========================================================================
s2 = wb.create_sheet("Retro Evidence")
darken_sheet(s2)

s2["A1"] = "Retro Evidence — Ubiquify all-time (pre-GigRadar + GigRadar-era hires)"
s2["A1"].font = fg_bold(FG, 16)

row = 3
s2.cell(row=row, column=1, value="Team joined GigRadar").font = fg_bold(FG_DIM)
s2.cell(row=row, column=2, value=phase1["team"]["joined_effective"][:10]).font = fg(FG)
row += 1
s2.cell(row=row, column=1, value="Subscription plan").font = fg_bold(FG_DIM)
s2.cell(row=row, column=2, value=phase1["team"].get("subscription_plan") or "—").font = fg(FG)
row += 1
s2.cell(row=row, column=1, value="Scanners (total / active)").font = fg_bold(FG_DIM)
n_total = phase2a["subject_scanner_count"]["total"]
n_active = phase2a["subject_scanner_count"]["active"]
s2.cell(row=row, column=2, value=f"{n_total} / {n_active}").font = fg(FG)
row += 1

default_cl = phase1["team"].get("default_cover_letter_config") or {}
s2.cell(row=row, column=1, value="Default CL config").font = fg_bold(FG_DIM)
s2.cell(row=row, column=2, value=json.dumps(default_cl)).font = fg_mono(FG)
row += 2

# Hired proposals timeline
s2.cell(row=row, column=1, value=f"HIRED proposals (n={len(phase1.get('hired_proposals') or [])})").font = fg_bold(FG, 14)
s2.cell(row=row, column=1).fill = fill(HEADER_BG)
fill_row_bg(s2, row, 1, 8, HEADER_BG)
row += 1
hired_headers = ["Hire Date", "Job Title", "Cohort", "Scanner", "Algorithm", "Connects", "Invite?", "CL Excerpt"]
for i, h in enumerate(hired_headers, 1):
    c = s2.cell(row=row, column=i, value=h)
    c.font = fg_bold(FG_DIM)
    c.fill = fill(HEADER_BG)
row += 1

for hp in (phase1.get("hired_proposals") or []):
    s2.cell(row=row, column=1, value=(hp.get("hire_ts_modified") or "")[:10]).font = fg(FG)
    s2.cell(row=row, column=2, value=(hp.get("job_title") or "")[:80]).font = fg(FG)
    s2.cell(row=row, column=3, value=hp.get("cohort")).font = fg(FG)
    s2.cell(row=row, column=4, value=hp.get("scanner_name") or "—").font = fg(FG)
    s2.cell(row=row, column=5, value=hp.get("algorithm_signature") or "—").font = fg_mono(FG)
    s2.cell(row=row, column=6, value=hp.get("connects_ladder")).font = fg(FG)
    s2.cell(row=row, column=7, value="YES" if hp.get("is_invite") else "no").font = fg(WIN_FG if hp.get("is_invite") else FG_DIM)
    cl = (hp.get("cover_letter_opp") or hp.get("cover_letter_proposal") or "")[:300]
    c = s2.cell(row=row, column=8, value=cl)
    c.font = fg_mono(FG)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    fill_row_bg(s2, row, 1, 8, BG_ALT)
    s2.row_dimensions[row].height = 42
    row += 1

fill_sheet_bg(s2, max_row=row+5, max_col=10)
set_col_widths(s2, [14, 42, 10, 24, 14, 10, 10, 70])


# =========================================================================
# Sheet 3 — Competitive Deep-Dive (v3: side-by-side competitor | Ubiquify analogue)
# Layout (14 cols, A..N):
#   LEFT pane  (A..F)  = competitor card
#   G                   = gutter
#   RIGHT pane (H..M)   = Ubiquify WIN or LOSS analogue
#   N                   = AI reasoning column (paragraphs)
# =========================================================================
s3 = wb.create_sheet("Competitive Deep-Dive")
darken_sheet(s3)

# helpers
def _fmt_money(v):
    if v is None: return "—"
    try:
        v = float(v)
        if v >= 1_000_000: return f"${v/1_000_000:.2f}M"
        if v >= 1_000: return f"${v/1_000:.1f}k"
        return f"${v:,.0f}"
    except Exception:
        return str(v)

def _clean(s, n=None):
    if s is None: return ""
    s = str(s).replace("\xa0", " ").strip()
    if n: s = s[:n]
    return s

LEFT_COLS = range(1, 7)   # A..F
RIGHT_COLS = range(8, 13)  # H..L (5 cols)
REASON_COL = 14            # N
GUT_COL = 7                # G

def pane_bg(ws, r, bg=BG_ALT):
    for c in list(LEFT_COLS) + list(RIGHT_COLS) + [REASON_COL]:
        ws.cell(row=r, column=c).fill = fill(bg)
    ws.cell(row=r, column=GUT_COL).fill = fill(BG)

def v_align_top(ws, r):
    for c in list(LEFT_COLS) + list(RIGHT_COLS) + [REASON_COL, GUT_COL]:
        cell = ws.cell(row=r, column=c)
        prev = cell.alignment
        cell.alignment = Alignment(
            horizontal=prev.horizontal, vertical="top", wrap_text=prev.wrap_text)

s3["A1"] = "Competitive Deep-Dive — Top-5 Competitor Wins vs. Ubiquify Analogues"
s3["A1"].font = fg_bold(FG, 16)
s3["A2"] = ("Left pane: each of the top-10 KNN competitors' 5 most-recent wins in Ubiquify's job neighborhoods. "
            "Right pane: Ubiquify's best title-matched WIN and LOSS for each. Column N: pattern observations per pair. "
            "Row colors: green = Ubiquify WIN analogue · red = Ubiquify LOSS analogue · gold = GigRadar template.")
s3["A2"].font = fg(FG_DIM)
s3["A2"].alignment = Alignment(wrap_text=True, vertical="top")
s3.merge_cells("A2:N2")
s3.row_dimensions[2].height = 48

row = 4

# ---- Cohort compare (4 rows — includes Ubiquify highlighted) ----
s3.cell(row=row, column=1,
    value="Cohort Compare (focus 2026-03-23..2026-04-22, ≥100 sent for 2a cohorts / ≥50 for KNN)"
).font = fg_bold(FG, 13)
fill_row_bg(s3, row, 1, 14, HEADER_BG)
s3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
row += 1

hdrs = ["Cohort", "N Teams", "Reply Rate (p25 / med / p75 / p90)",
        "$/Reply (p25 / med / p75 / p90)", "Ubiquify RR / $/Reply",
        "Percentile Rank"]
for i, h in enumerate(hdrs, 1):
    c = s3.cell(row=row, column=i, value=h)
    c.font = fg_bold(FG_DIM)
    c.fill = fill(HEADER_BG)
    c.alignment = Alignment(vertical="top", wrap_text=True)
fill_row_bg(s3, row, 7, 14, HEADER_BG)
row += 1

cohort_rows = [
    ("Inferred peers (scanner-keyword match)", phase2a["inferred_cohort"]),
    ("Broad platform (any ≥100 sent)", phase2a["broad_cohort"]),
    ("KNN-neighbor jobs (Ubiquify territorial peers)", phase2b["knn_cohort_summary"]),
]
# Geo cohort — added v0.4: anchors a country-matched cohort off the subject
# agency's primary location. Useful because price expectations + buyer-side
# language correlate with freelancer country.
geo_c = phase2a.get("geo_cohort")
if geo_c:
    geo_country = phase2a.get("geo_cohort_country") or "?"
    cohort_rows.append((f"Geo peers (agency primary country = {geo_country})", geo_c))
for label, d in cohort_rows:
    rrs = d["reply_rate_stats"]
    cprs = d["cost_per_reply_stats"]
    sp = d["subject_position"]
    is_knn = "KNN" in label
    s3.cell(row=row, column=1, value=label).font = fg_bold(WIN_FG if is_knn else FG)
    s3.cell(row=row, column=2, value=d["qualified_count"]).font = fg(FG)
    s3.cell(row=row, column=3, value=f"{rrs['p25']*100:.2f}% / {rrs['median']*100:.2f}% / {rrs['p75']*100:.2f}% / {rrs['p90']*100:.2f}%").font = fg(FG)
    s3.cell(row=row, column=4, value=f"${cprs['p25']:.2f} / ${cprs['median']:.2f} / ${cprs['p75']:.2f} / ${cprs['p90']:.2f}").font = fg(FG)
    s3.cell(row=row, column=5, value=f"{sp['reply_rate']*100:.2f}%  |  ${sp['cost_per_reply']:.2f}").font = fg(FG)
    rr_p = sp['reply_rate_percentile']
    cpr_p = sp['cost_per_reply_percentile_lower_better']
    color = WIN_FG if (rr_p >= 60 and cpr_p >= 60) else (OKAY_FG if rr_p >= 40 else CRIT_FG)
    s3.cell(row=row, column=6, value=f"RR P{rr_p:.0f}  |  $/reply P{cpr_p:.0f}").font = fg_bold(color)
    for cc in range(1, 7):
        s3.cell(row=row, column=cc).fill = fill(BG_ALT)
        s3.cell(row=row, column=cc).alignment = Alignment(vertical="top", wrap_text=True)
    fill_row_bg(s3, row, 7, 14, BG)
    s3.row_dimensions[row].height = 30
    row += 1

# ---- Ubiquify self-row (highlighted) ----
ub_fw = phase1["focus_window"]["auto_bidder"]
ub_rr = ub_fw["reply_rate"]
ub_cpr = ub_fw["cost_per_reply_usd"]
# Percentile ranks vs each cohort
knn_sp = phase2b["knn_cohort_summary"]["subject_position"]
inf_sp = phase2a["inferred_cohort"]["subject_position"]
brd_sp = phase2a["broad_cohort"]["subject_position"]
geo_sp = (phase2a.get("geo_cohort") or {}).get("subject_position") or {}
s3.cell(row=row, column=1, value="◆ UBIQUIFY (you) — subject team").font = fg_bold(UB_HL_FG)
s3.cell(row=row, column=2, value=1).font = fg_bold(UB_HL_FG)
s3.cell(row=row, column=3, value=f"{ub_rr*100:.2f}% — sent {ub_fw['sent']}, replies {ub_fw['replied']}").font = fg(UB_HL_FG)
s3.cell(row=row, column=4, value=f"${ub_cpr:.2f} — connects ${ub_fw['connects_spend_usd']:.0f}").font = fg(UB_HL_FG)
s3.cell(row=row, column=5, value=f"{ub_rr*100:.2f}%  |  ${ub_cpr:.2f}").font = fg_bold(UB_HL_FG)
percentile_parts = [
    f"vs Inferred P{inf_sp['reply_rate_percentile']:.0f}/P{inf_sp['cost_per_reply_percentile_lower_better']:.0f}",
    f"vs Broad P{brd_sp['reply_rate_percentile']:.0f}/P{brd_sp['cost_per_reply_percentile_lower_better']:.0f}",
    f"vs KNN P{knn_sp['reply_rate_percentile']:.0f}/P{knn_sp['cost_per_reply_percentile_lower_better']:.0f}",
]
if geo_sp and geo_sp.get("reply_rate_percentile") is not None:
    percentile_parts.append(
        f"vs Geo P{geo_sp['reply_rate_percentile']:.0f}/P{geo_sp.get('cost_per_reply_percentile_lower_better') or 0:.0f}"
    )
s3.cell(row=row, column=6, value="  ·  ".join(percentile_parts)).font = fg_bold(UB_HL_FG)
for cc in range(1, 7):
    s3.cell(row=row, column=cc).fill = fill(UB_HL_BG)
    s3.cell(row=row, column=cc).alignment = Alignment(vertical="top", wrap_text=True)
fill_row_bg(s3, row, 7, 14, BG)
s3.row_dimensions[row].height = 38
row += 2

competitors = phase2b.get("competitors_top10", [])
c2c_lookup = {c["team_id"]: c for c in phase2c.get("competitors", [])}

# ─── Sheet-3 DISPLAY ORDER ────────────────────────────────────────────────
# SELECTION lives in phase2b (top-10 by neighborhood overlap ≥2 — real
# competitors on shared job types). DISPLAY ORDER on Sheet 3 uses the
# composite performance score so readers see the highest-leverage
# competitors first — highest reply rate + hire rate + volume closers.
#
#   score = reply_rate + 2·hire_rate + min(hires/10, 0.5)
#
# Same formula as the exec-summary tactic ranking (single source of truth);
# missing focus_metrics → score 0 (sort to bottom).
def _comp_display_score(comp):
    fm = comp.get("focus_metrics") or {}
    rr = fm.get("reply_rate") or 0.0
    hr = fm.get("hire_rate") or 0.0
    h = fm.get("hired") or 0
    return rr + 2.0 * hr + min(h / 10.0, 0.5)

competitors = sorted(competitors, key=lambda c: -_comp_display_score(c))

# =========================================================================
# Profile Side-by-Side block — NEW (v0.4)
# For each top-10 competitor: compare their agency + primary freelancer
# profile with Ubiquify's. All fields the reader needs to see *who is winning
# and why their positioning works* — avatars (embedded as images), rates,
# earnings, JSS, skills (top 10), services, work history, portfolios.
# Sorted in the same order as the per-competitor cards below (by reply rate).
# =========================================================================
s3.cell(row=row, column=1,
    value="Freelancer & Agency Profile Comparison — Competitor vs Ubiquify  (top-10 ranked by focus-window reply rate)"
).font = fg_bold(FG, 13)
for cc in range(1, 15):
    s3.cell(row=row, column=cc).fill = fill(HEADER_BG)
s3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
s3.row_dimensions[row].height = 26
row += 1
s3.cell(row=row, column=1,
    value="Cover-letter copy alone doesn't explain wins — a $95/hr 'Digital Marketing Automation expert' with 100% JSS wins differently than a $60/hr 'Full-Stack Generalist'. Read profile positioning first, then scroll down to the bid-by-bid analysis."
).font = fg(FG_DIM)
s3.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
s3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
s3.row_dimensions[row].height = 32
row += 1

# Ubiquify subject profile (lifted once — used as the right-pane reference)
subj_profile = phase2b.get("subject_profile") or {}
subj_agency = subj_profile.get("agency") or {}
subj_fl_top = subj_profile.get("main_freelancer") or {}
subj_fl = (subj_fl_top.get("data") or {}) if isinstance(subj_fl_top, dict) else {}

def _profile_stats_line(agency):
    stats = (agency or {}).get("stats") or {}
    parts = []
    te = stats.get("totalEarning")
    if te: parts.append(f"${te:,.0f} earned")
    jss = stats.get("jobSuccessScore")
    if isinstance(jss, (int, float)) and jss > 0:
        parts.append(f"{jss*100:.0f}% JSS")
    tj = stats.get("totalJobs")
    if tj: parts.append(f"{tj} jobs")
    locs = (agency or {}).get("locations") or []
    loc_str = ", ".join(f"{(l or {}).get('city','')}, {(l or {}).get('country','')}".strip(", ")
                         for l in locs[:2])
    if loc_str: parts.append(loc_str)
    return " · ".join(parts) or "—"

def _fl_stats_line(fl_data):
    if not fl_data: return "—"
    st = fl_data.get("stats") or {}
    parts = []
    te = st.get("total_earnings")
    if te: parts.append(f"${te:,.0f} earned")
    tj = st.get("total_jobs")
    if tj: parts.append(f"{tj} jobs")
    jss = st.get("job_success_score")
    if isinstance(jss, (int, float)) and jss > 0:
        parts.append(f"{jss*100:.0f}% JSS" if jss <= 1 else f"{jss:.0f}% JSS")
    tier = fl_data.get("contractor_tier")
    if tier: parts.append(f"tier {tier}")
    return " · ".join(parts) or "—"

def _history_line(fl_data, n=3):
    eh = (fl_data or {}).get("employment_history") or []
    if not eh: return ""
    lines = []
    for e in eh[:n]:
        ln = f"• {e.get('title','')} @ {e.get('company','')}  ({e.get('start','')[:4]}-{(e.get('end') or '')[:4] if e.get('end')!='present' else 'now'})"
        if e.get("summary"):
            ln += f" — {e.get('summary','')[:150]}"
        lines.append(ln)
    return "\n".join(lines) if lines else ""

def _portfolio_line(agency, n=4):
    ports = (agency or {}).get("portfolios") or []
    if not ports: return ""
    lines = []
    for p in ports[:n]:
        t = (p or {}).get("title") or "—"
        d = ((p or {}).get("description") or "")[:150]
        lines.append(f"• {t}  {('— ' + d) if d else ''}")
    return "\n".join(lines)

def _skills_line(x, n=12):
    sk = (x or {}).get("skills") or []
    return ", ".join(sk[:n])

def _services_line(agency, n=5):
    svc = (agency or {}).get("services") or []
    names = []
    for s in svc[:n]:
        n_ = (s or {}).get("name") or (s or {}).get("title")
        if n_: names.append(n_)
    return " · ".join(names)

def render_profile_card(r, col_start, col_end, side_label, side_color, comp_or_subj):
    """Render one profile card — AVATAR row (image) + NAME row + STATS +
    FL ROW + DESC + SKILLS + SERVICES + HISTORY + PORTFOLIOS. Returns final row."""
    agency = comp_or_subj.get("agency") or {}
    fl_top = comp_or_subj.get("freelancer") or comp_or_subj.get("main_freelancer") or {}
    fl_data = (fl_top.get("data") or {}) if isinstance(fl_top, dict) else {}

    # AVATAR row — agency photo + freelancer portrait side by side inside
    # the card. Insert images via openpyxl Image anchor (agency in col+1, fl
    # in col+2 when width allows). Row height set to ~80px so images fit.
    label_col = col_start
    s3.cell(row=r, column=label_col, value=side_label).font = fg_bold(side_color, 11)
    s3.cell(row=r, column=label_col).alignment = Alignment(vertical="top", wrap_text=True)
    # Try to anchor agency logo next to label
    ag_img = _download_and_size_image(agency.get("photoUrl"), size=(70, 70))
    fl_img = _download_and_size_image((fl_data or {}).get("portrait_url"), size=(70, 70))
    if ag_img is not None:
        anchor_ag = f"{get_column_letter(col_start + 1)}{r}"
        s3.add_image(ag_img, anchor_ag)
    if fl_img is not None:
        anchor_fl = f"{get_column_letter(col_start + 2)}{r}"
        s3.add_image(fl_img, anchor_fl)
    # Name/Link labels (to the right of avatars) in remaining columns
    _name_cell = s3.cell(row=r, column=col_start + 3,
                         value=f"{agency.get('name') or '—'}  /  {fl_data.get('general_name') or '—'}")
    _name_cell.font = fg_bold(LINK_FG, 11)
    _name_cell.alignment = Alignment(vertical="top", wrap_text=True)
    if agency.get("url"):
        _name_cell.hyperlink = Hyperlink(ref=_name_cell.coordinate, target=agency.get("url"),
                                          tooltip="agency on Upwork")
    if col_end > col_start + 3:
        s3.merge_cells(start_row=r, start_column=col_start + 3, end_row=r, end_column=col_end)
    s3.row_dimensions[r].height = 80
    pane_bg(s3, r, BG_ALT)
    r += 1

    # STATS row (agency + freelancer)
    stats_line = f"AGENCY  {_profile_stats_line(agency)}\nFREELANCER  {_fl_stats_line(fl_data)}"
    lc = s3.cell(row=r, column=col_start, value="STATS")
    lc.font = fg_bold(FG_DIM, 9); lc.alignment = Alignment(vertical="top", wrap_text=True)
    sc = s3.cell(row=r, column=col_start + 1, value=stats_line)
    sc.font = fg(FG); sc.alignment = Alignment(vertical="top", wrap_text=True)
    s3.merge_cells(start_row=r, start_column=col_start + 1, end_row=r, end_column=col_end)
    s3.row_dimensions[r].height = 34
    pane_bg(s3, r, BG_ALT)
    r += 1

    # FL TITLE (specialized-profile or general)
    title_text = (fl_data.get("specialized") or {}).get("title") or fl_data.get("general_title") or "—"
    rate = fl_data.get("general_hourly_rate")
    rate_txt = f" · ${rate}/hr" if rate else ""
    lc = s3.cell(row=r, column=col_start, value="FL TITLE")
    lc.font = fg_bold(FG_DIM, 9); lc.alignment = Alignment(vertical="top", wrap_text=True)
    tc = s3.cell(row=r, column=col_start + 1, value=f"{title_text}{rate_txt}")
    tc.font = fg_bold(LINK_FG, 11); tc.alignment = Alignment(vertical="top", wrap_text=True)
    if fl_top.get("url"):
        tc.hyperlink = Hyperlink(ref=tc.coordinate, target=fl_top.get("url"),
                                  tooltip="freelancer on Upwork")
    s3.merge_cells(start_row=r, start_column=col_start + 1, end_row=r, end_column=col_end)
    s3.row_dimensions[r].height = 20
    pane_bg(s3, r, BG_ALT)
    r += 1

    # FL DESC (first 400 chars of nested profile description)
    desc = fl_data.get("general_description") or ""
    lc = s3.cell(row=r, column=col_start, value="FL DESC")
    lc.font = fg_bold(FG_DIM, 9); lc.alignment = Alignment(vertical="top", wrap_text=True)
    dc = s3.cell(row=r, column=col_start + 1, value=desc[:800] or "—")
    dc.font = fg(FG); dc.alignment = Alignment(vertical="top", wrap_text=True)
    s3.merge_cells(start_row=r, start_column=col_start + 1, end_row=r, end_column=col_end)
    s3.row_dimensions[r].height = max(60, estimate_cl_rows_h(desc[:800], cols_width_chars=(col_end - col_start) * 10) * 13)
    pane_bg(s3, r, BG_ALT)
    r += 1

    # AGENCY summary/overview
    ov = agency.get("overview") or agency.get("summary_or_description") or ""
    if ov:
        lc = s3.cell(row=r, column=col_start, value="AGENCY DESC")
        lc.font = fg_bold(FG_DIM, 9); lc.alignment = Alignment(vertical="top", wrap_text=True)
        ac = s3.cell(row=r, column=col_start + 1, value=ov[:700])
        ac.font = fg(FG); ac.alignment = Alignment(vertical="top", wrap_text=True)
        s3.merge_cells(start_row=r, start_column=col_start + 1, end_row=r, end_column=col_end)
        s3.row_dimensions[r].height = max(45, estimate_cl_rows_h(ov[:700], cols_width_chars=(col_end - col_start) * 10) * 13)
        pane_bg(s3, r, BG_ALT)
        r += 1

    # SKILLS (FL) + SERVICES (AG)
    sk = _skills_line(fl_data)
    svc = _services_line(agency)
    mixed = []
    if sk: mixed.append(f"FL skills: {sk}")
    if svc: mixed.append(f"AG services: {svc}")
    if (agency or {}).get("skills"):
        mixed.append("AG skills: " + ", ".join(((agency.get("skills") or [])[:10])))
    if mixed:
        lc = s3.cell(row=r, column=col_start, value="SKILLS")
        lc.font = fg_bold(FG_DIM, 9); lc.alignment = Alignment(vertical="top", wrap_text=True)
        skc = s3.cell(row=r, column=col_start + 1, value="\n".join(mixed))
        skc.font = fg(OKAY_FG); skc.alignment = Alignment(vertical="top", wrap_text=True)
        s3.merge_cells(start_row=r, start_column=col_start + 1, end_row=r, end_column=col_end)
        s3.row_dimensions[r].height = 54
        pane_bg(s3, r, BG_ALT)
        r += 1

    # WORK HISTORY (FL employment_history top-3)
    hist = _history_line(fl_data, n=3)
    if hist:
        lc = s3.cell(row=r, column=col_start, value="HISTORY")
        lc.font = fg_bold(FG_DIM, 9); lc.alignment = Alignment(vertical="top", wrap_text=True)
        hc = s3.cell(row=r, column=col_start + 1, value=hist)
        hc.font = fg(FG); hc.alignment = Alignment(vertical="top", wrap_text=True)
        s3.merge_cells(start_row=r, start_column=col_start + 1, end_row=r, end_column=col_end)
        s3.row_dimensions[r].height = max(45, estimate_cl_rows_h(hist, cols_width_chars=(col_end - col_start) * 10) * 13)
        pane_bg(s3, r, BG_ALT)
        r += 1

    # PORTFOLIOS (AG portfolio top-4)
    port = _portfolio_line(agency, n=4)
    if port:
        lc = s3.cell(row=r, column=col_start, value="PORTFOLIO")
        lc.font = fg_bold(FG_DIM, 9); lc.alignment = Alignment(vertical="top", wrap_text=True)
        pc = s3.cell(row=r, column=col_start + 1, value=port)
        pc.font = fg(FG); pc.alignment = Alignment(vertical="top", wrap_text=True)
        s3.merge_cells(start_row=r, start_column=col_start + 1, end_row=r, end_column=col_end)
        s3.row_dimensions[r].height = max(50, estimate_cl_rows_h(port, cols_width_chars=(col_end - col_start) * 10) * 13)
        pane_bg(s3, r, BG_ALT)
        r += 1

    return r


def estimate_cl_rows_h(text, cols_width_chars):
    """Stub — actual defined later in the file; rely on forward lookup at call time."""
    if not text: return 2
    ln = len(text)
    return max(2, min(18, ln // cols_width_chars + 2))


# Render profile block per competitor
for prof_idx, comp in enumerate(competitors[:10], 1):
    tid = comp.get("team_id")
    team_name = comp.get("team_name") or tid[:12]
    fm = comp.get("focus_metrics") or {}
    rr = fm.get("reply_rate")
    rr_txt = f"{rr*100:.1f}%" if isinstance(rr, (int, float)) else "—"
    cpr = fm.get("cost_per_reply")
    cpr_txt = f"${cpr:.0f}/reply" if isinstance(cpr, (int, float)) else "—"
    sent_txt = f"{fm.get('sent') or '—'} sent"
    hires_txt = f"{fm.get('hired') or 0} hires"
    hdr = (f"  #{prof_idx}  ·  {team_name}  ·  "
           f"{rr_txt} RR  ·  {cpr_txt}  ·  {sent_txt}  ·  {hires_txt}")
    c = s3.cell(row=row, column=1, value=hdr)
    c.font = fg_bold(WIN_FG if rr and rr >= 0.08 else FG, 12)
    c.alignment = Alignment(vertical="center")
    for cc in range(1, 15):
        s3.cell(row=row, column=cc).fill = fill(HEADER_BG)
    s3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    s3.row_dimensions[row].height = 24
    row += 1

    # Render two profile cards side by side: LEFT = competitor, RIGHT = Ubiquify
    block_start = row
    left_end = render_profile_card(row, 1, 6, "COMPETITOR", LINK_FG, comp)
    right_end = render_profile_card(row, 8, 12, "UBIQUIFY", UB_HL_FG,
                                     {"agency": subj_agency, "main_freelancer": subj_fl_top})
    # Align panes (pad shorter with empty rows)
    final_row = max(left_end, right_end)
    for rr_pad in range(min(left_end, right_end), final_row):
        pane_bg(s3, rr_pad, BG_ALT)
    row = final_row

    # AI profile positioning on col N — merged vertically across card rows.
    # Some subagents emit a dict for this field (schema drift from the prompt
    # examples); coerce to string before rendering.
    c2c = c2c_lookup.get(tid) or {}
    pp_raw = c2c.get("ai_profile_positioning")
    if isinstance(pp_raw, dict):
        pp_text = "\n".join(f"{k}: {v}" for k, v in pp_raw.items())
    elif isinstance(pp_raw, list):
        pp_text = "\n".join(str(x) for x in pp_raw)
    else:
        pp_text = str(pp_raw or "")
    if pp_text:
        pc = s3.cell(row=block_start, column=REASON_COL, value="PROFILE DELTA\n\n" + pp_text)
        pc.font = fg(FG); pc.alignment = Alignment(wrap_text=True, vertical="top")
        if final_row > block_start:
            s3.merge_cells(start_row=block_start, start_column=REASON_COL,
                           end_row=final_row, end_column=REASON_COL)

    # Separator row
    for cc in range(1, 15):
        s3.cell(row=row, column=cc).fill = fill(BG)
    s3.row_dimensions[row].height = 10
    row += 1

# ---- Section header: bid-by-bid cards below ----
s3.cell(row=row, column=1,
    value="Top-10 Direct Competitors × Top-5 Wins — Side-by-Side with Ubiquify's Matched Win / Loss"
).font = fg_bold(FG, 13)
for cc in range(1, 15):
    s3.cell(row=row, column=cc).fill = fill(HEADER_BG)
s3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
s3.row_dimensions[row].height = 24
row += 1
s3.cell(row=row, column=1,
    value="Matching = ES metajob KNN vector cosine + title Jaccard fallback. "
          "See COMPETITIVE_PLAYBOOK.md for the reading guide and reasoning framework."
).font = fg(FG_DIM)
s3.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
s3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
s3.row_dimensions[row].height = 28
row += 2

def render_cl_block(ws, r, side, label, meta_text, cl_text, cl_color=FG):
    """Render a CL block on left (side='L') or right (side='R') pane.
    Row height sized to CL length.
    """
    if side == "L":
        lbl_col = 1
        meta_col = 2
        body_cols = (3, 6)   # merge C:F
    else:  # R
        lbl_col = 8
        meta_col = 9
        body_cols = (10, 12)  # merge J:L
    ws.cell(row=r, column=lbl_col, value=label).font = fg_bold(FG_DIM, 9)
    ws.cell(row=r, column=lbl_col).alignment = Alignment(vertical="top", wrap_text=True)
    ws.cell(row=r, column=meta_col, value=meta_text).font = fg(FG_DIM)
    ws.cell(row=r, column=meta_col).alignment = Alignment(wrap_text=True, vertical="top")
    bc = ws.cell(row=r, column=body_cols[0], value=cl_text or "")
    bc.font = fg_mono(cl_color, size=9)
    bc.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=body_cols[0], end_row=r, end_column=body_cols[1])

def estimate_cl_rows_h(cl_text, cols_width_chars=60):
    ln = len(cl_text or "")
    return max(3, min(22, ln // cols_width_chars + 4))

for idx, comp in enumerate(competitors, 1):
    agency = comp.get("agency") or {}
    flancer = comp.get("freelancer") or {}
    flancer_data = (flancer.get("data") or {}) if isinstance(flancer, dict) else {}
    gbid = comp.get("gigradar_bid") or {}

    is_gigradar = gbid.get("is_gigradar_bid")
    algo_name = gbid.get("algorithm_name") or "—"
    source_tag = f"GigRadar · {algo_name}" if is_gigradar else "Direct bid (non-GigRadar)"
    wins = comp.get("wins_in_neighborhoods")
    team_name = comp.get("team_name") or comp.get("team_id", "")[:12]

    # 1) Header (A:N merged) — includes focus-window reply rate to show WHY this
    # competitor is ranked here, not just their name.
    c2c = c2c_lookup.get(comp.get("team_id")) or {}
    fm = comp.get("focus_metrics") or {}
    _rr = fm.get("reply_rate")
    _rr_txt = f"{_rr*100:.1f}% RR" if isinstance(_rr, (int, float)) else "—"
    _cpr = fm.get("cost_per_reply")
    _cpr_txt = f"${_cpr:.0f}/reply" if isinstance(_cpr, (int, float)) else "—"
    hdr_text = (f"#{idx}  ·  {team_name}  ·  {_rr_txt}  ·  {_cpr_txt}  ·  "
                f"{fm.get('sent') or '—'} sent  ·  {fm.get('hired') or 0} hires  "
                f"·  {wins} win(s) in Ubiquify neighborhoods  ·  {source_tag}")
    c = s3.cell(row=row, column=1, value=hdr_text)
    c.font = fg_bold(WIN_FG if is_gigradar else FG, size=13)
    c.alignment = Alignment(vertical="center")
    for cc in range(1, 15):
        s3.cell(row=row, column=cc).fill = fill(HEADER_BG)
    s3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    s3.row_dimensions[row].height = 26
    row += 1

    # 2) COMPETITOR FORMULA — STANDALONE PROMINENT ROW (v0.4).
    # Promoted out of col N (where it was buried) into a full-width,
    # high-contrast band. This is the single most actionable cell per competitor
    # — the reader should see the formula BEFORE scrolling through paired bids.
    ai_summary = c2c.get("ai_summary") or c2c.get("summary_reasoning") or ""
    ai_tactics = c2c.get("ai_tactics") or []
    formula_body = "▸ COMPETITOR FORMULA\n" + (ai_summary or "—")
    if ai_tactics:
        formula_body += "\n\n▸ TACTICS FOR UBIQUIFY TO COPY"
        for t in ai_tactics:
            formula_body += f"\n  • {t}"
    fc = s3.cell(row=row, column=1, value=formula_body)
    fc.font = fg_bold(UB_HL_FG, 11)
    fc.alignment = Alignment(wrap_text=True, vertical="top")
    # Fill with a distinctive highlighted band so the row reads as a callout.
    for cc in range(1, 15):
        s3.cell(row=row, column=cc).fill = fill(UB_HL_BG)
    s3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    formula_lines = estimate_cl_rows_h(formula_body, cols_width_chars=160)
    s3.row_dimensions[row].height = max(60, formula_lines * 14)
    row += 1

    # 3) UBIQUIFY analogue header row — just a divider line before the bid pairs.
    s3.cell(row=row, column=1, value="").font = fg(FG_DIM, 9)
    s3.cell(row=row, column=8, value="UBIQUIFY").font = fg_bold(UB_HL_FG, 10)
    s3.cell(row=row, column=9, value="matched win + loss analogues per competitor win (right pane)").font = fg(FG_DIM)
    s3.merge_cells(start_row=row, start_column=9, end_row=row, end_column=12)
    pane_bg(s3, row, BG_ALT)
    for cc in range(1, 15):
        s3.cell(row=row, column=cc).alignment = Alignment(vertical="top", wrap_text=True)
    s3.row_dimensions[row].height = 22
    row += 1

    # 4) Per-win pairs — SYMMETRIC 3-CARD RENDER
    # Each pair has up to three bid cards laid out side-by-side + vertically:
    #
    #   LEFT pane (A-F):     RIGHT pane upper (H-L):   | col N (full height)
    #   COMP WIN card         UB WIN card               | AI REASONING:
    #                                                   |   WHAT WORKED FOR THEM
    #   (LEFT continues       RIGHT pane lower:         |   WHAT UBIQUIFY DID
    #    w/ faded note)       UB LOSS card              |   TACTIC TO COPY
    #
    # Each card uses the same 6-section shape so the reader can eyeball deltas
    # field-by-field:
    #     ① JOB      — title (link) + JD excerpt
    #     ② CLIENT   — company · country · $spent · feedback · hire rate
    #     ③ BID      — amount + connects
    #     ④ FL       — name (link) · $rate · location \n title \n description
    #     ⑤ CL       — full rendered cover letter (monospace)
    #     ⑥ TMPL     — frozen bid-time template (from opp.originalStrategy)

    def _match_badge(m):
        """Returns (badge_text, color) for a matched UB proposal, given its
        match_mode and cosine similarity score from phase2c vector matching."""
        if not m:
            return ("—", FG_DIM)
        mm = m.get("match_mode") or ""
        mc = m.get("match_cosine")
        if mm == "cosine_tight":
            return (f"✓ {int(mc*100)}% sim", WIN_FG)
        if mm == "cosine_loose":
            return (f"~ {int(mc*100)}% sim (loose)", OKAY_FG)
        if mm == "token_fallback":
            j = m.get("match_jaccard") or 0.0
            return (f"⚠ token-jaccard={j}", OKAY_FG)
        return (mm, FG_DIM)

    def _client_line(comp_dict):
        """Build the one-line CLIENT summary used in every card's CLIENT row."""
        parts = []
        co = comp_dict.get("client_company")
        if co: parts.append(_clean(co, 30))
        country = comp_dict.get("client_country")
        if country: parts.append(country)
        spent = comp_dict.get("client_total_spent")
        if spent: parts.append(f"spent {_fmt_money(spent)}")
        fb = comp_dict.get("client_feedback_score")
        if isinstance(fb, (int, float)) and fb > 0:
            parts.append(f"{fb:.1f}★")
        hr = comp_dict.get("client_hire_rate")
        if isinstance(hr, (int, float)) and hr > 0:
            parts.append(f"{hr*100:.0f}% hire rate" if hr <= 1 else f"{hr:.0f}% hire rate")
        return " · ".join(parts) or "—"

    def _bid_line(bid_amt, connects, is_invite=False):
        parts = [f"bid: {_fmt_money(bid_amt)}"]
        parts.append(f"{connects or '—'} connects")
        if is_invite:
            parts.append("📩 via invite")
        return " · ".join(parts)

    def _fl_block(fr):
        """Render the 3-line freelancer summary — name/rate/loc on line 1,
        title on line 2, description excerpt on line 3."""
        if not fr or not isinstance(fr, dict):
            return None, None
        name = _clean(fr.get("name"), 50) or "—"
        rate = fr.get("hourly_rate")
        rate_txt = f"${rate}/hr" if rate else ""
        loc = _clean(fr.get("location"), 30)
        line1_parts = [name]
        if rate_txt: line1_parts.append(rate_txt)
        if loc: line1_parts.append(loc)
        line1 = " · ".join(line1_parts)
        title = _clean(fr.get("title"), 100)
        desc = _clean(fr.get("description"), 220)
        body_parts = []
        if title: body_parts.append(title)
        if desc: body_parts.append(desc)
        body = "\n".join(body_parts)
        url = fr.get("url")
        return (line1, body, url)

    def _lookup_frozen_template_for_comp(pid):
        """Pull the bid-time template snapshot for a competitor win by
        joining back into phase2b.winning_proposals_top5 by proposal_id."""
        for w_raw in (comp.get("winning_proposals_top5") or []):
            if w_raw.get("proposal_id") == pid:
                gb = w_raw.get("gigradar_bid") or {}
                return gb.get("cl_template_used"), gb.get("scanner_name"), gb.get("match_percentage")
        return None, None, None

    def render_card_row(r, col_start, col_end, label, label_color,
                        header_text, header_color,
                        body_text=None, body_color=FG_DIM, body_mono=False,
                        url=None, row_height=None):
        """Render ONE data row of a bid card — label in col_start, merged
        header+body cell from col_start+1..col_end. Returns nothing; caller
        increments the row counter."""
        lc = s3.cell(row=r, column=col_start, value=label)
        lc.font = fg_bold(label_color, 9)
        lc.alignment = Alignment(vertical="top", wrap_text=True)
        # Merged content area
        hc = s3.cell(row=r, column=col_start + 1, value=header_text)
        hc.font = fg_bold(header_color, 10) if header_color else fg(FG)
        if url:
            hc.hyperlink = Hyperlink(ref=hc.coordinate, target=url, tooltip=url)
        hc.alignment = Alignment(vertical="top", wrap_text=True)
        if col_end > col_start + 1:
            s3.merge_cells(start_row=r, start_column=col_start + 1, end_row=r, end_column=col_end)
        if body_text:
            # add body on a separate line in the same cell via newline
            hc.value = (header_text or "") + ("\n" + body_text if header_text else body_text)
            if body_mono:
                hc.font = fg_mono(body_color, size=9)
        if row_height:
            # only grow, never shrink
            cur = s3.row_dimensions[r].height or 0
            s3.row_dimensions[r].height = max(cur, row_height)
        return r

    def render_bid_card(start_row, col_start, col_end, header_label, header_color,
                        card_data, cl_color=FG, match_badge_text=None, match_badge_color=None,
                        is_missing_reason=None):
        """Render one complete bid card (6 sections) into the rectangle
        [start_row..end_row] x [col_start..col_end]. Returns end_row (inclusive).
        card_data is a dict with keys:
          job_title, job_url, jd_excerpt,
          client_company/country/total_spent/feedback_score/hire_rate,
          bid_amount, connects_bid, is_invite,
          freelancer (dict), full_cl, cl_template_frozen.
        is_missing_reason: when present, writes a single row with that message
        and skips the rest of the card (used for "no comparable UB bid")."""
        r = start_row

        # ── JOB (header + excerpt in one cell) ──
        header_label_final = header_label
        if match_badge_text:
            header_label_final = f"{header_label}  {match_badge_text}"
        if is_missing_reason:
            render_card_row(r, col_start, col_end, "JOB", header_color,
                            is_missing_reason, CRIT_FG,
                            row_height=40)
            pane_bg(s3, r, BG_ALT)
            # header row label override
            s3.cell(row=r, column=col_start, value=header_label_final).font = fg_bold(match_badge_color or header_color, 10)
            return r

        job_title = _clean(card_data.get("job_title"), 140) or "—"
        jd_excerpt = _clean(card_data.get("jd_excerpt"), 400)
        # header label (UB WIN / UB LOSS / THEIR WIN) above; title in cell
        r_cell = s3.cell(row=r, column=col_start, value=header_label_final)
        r_cell.font = fg_bold(match_badge_color or header_color, 10)
        r_cell.alignment = Alignment(vertical="top", wrap_text=True)
        header_val = job_title
        if jd_excerpt:
            header_val = f"{job_title}\n{jd_excerpt}"
        h = s3.cell(row=r, column=col_start + 1, value=header_val)
        h.font = fg_bold(LINK_FG, 11)
        h.alignment = Alignment(vertical="top", wrap_text=True)
        if card_data.get("job_url"):
            h.hyperlink = Hyperlink(ref=h.coordinate, target=card_data.get("job_url"),
                                    tooltip=card_data.get("job_url"))
        s3.merge_cells(start_row=r, start_column=col_start + 1, end_row=r, end_column=col_end)
        jd_lines = estimate_cl_rows_h(header_val, cols_width_chars=(col_end - col_start) * 10)
        s3.row_dimensions[r].height = max(32, jd_lines * 14)
        pane_bg(s3, r, BG_ALT)
        r += 1

        # ── CLIENT ──
        render_card_row(r, col_start, col_end, "CLIENT", FG_DIM,
                        _client_line(card_data), FG, row_height=16)
        pane_bg(s3, r, BG_ALT)
        r += 1

        # ── BID ──
        render_card_row(r, col_start, col_end, "BID", FG_DIM,
                        _bid_line(card_data.get("bid_amount"), card_data.get("connects_bid"),
                                  is_invite=bool(card_data.get("is_invite"))),
                        WIN_FG, row_height=16)
        pane_bg(s3, r, BG_ALT)
        r += 1

        # ── FREELANCER ──
        fr = card_data.get("freelancer") or None
        fr_line1, fr_body, fr_url = _fl_block(fr) if fr else (None, None, None)
        # URL may come from phase4 (UB side) where freelancer is in `freelancer_profile`
        if not fr_line1 and card_data.get("freelancer_profile"):
            fr_line1, fr_body, fr_url = _fl_block(card_data.get("freelancer_profile"))
        if not fr_line1 and card_data.get("freelancer_name"):
            fr_line1 = card_data.get("freelancer_name")
        fl_cell_text = fr_line1 or "—"
        if fr_body:
            fl_cell_text = f"{fr_line1}\n{fr_body}"
        lc = s3.cell(row=r, column=col_start, value="FL")
        lc.font = fg_bold(FG_DIM, 9); lc.alignment = Alignment(vertical="top", wrap_text=True)
        fc = s3.cell(row=r, column=col_start + 1, value=fl_cell_text)
        fc.font = fg(FG); fc.alignment = Alignment(vertical="top", wrap_text=True)
        if fr_url:
            fc.hyperlink = Hyperlink(ref=fc.coordinate, target=fr_url, tooltip=fr_url)
        s3.merge_cells(start_row=r, start_column=col_start + 1, end_row=r, end_column=col_end)
        fl_lines = estimate_cl_rows_h(fl_cell_text, cols_width_chars=(col_end - col_start) * 10)
        s3.row_dimensions[r].height = max(36, fl_lines * 14)
        pane_bg(s3, r, BG_ALT)
        r += 1

        # ── CL (full rendered cover letter) ──
        cl_text = card_data.get("full_cl") or card_data.get("cl") or ""
        cl_len = card_data.get("cl_length") or len(cl_text)
        lc = s3.cell(row=r, column=col_start, value=f"CL\n{cl_len} ch")
        lc.font = fg_bold(FG_DIM, 9); lc.alignment = Alignment(vertical="top", wrap_text=True)
        cc = s3.cell(row=r, column=col_start + 1, value=cl_text)
        cc.font = fg_mono(cl_color, size=9); cc.alignment = Alignment(vertical="top", wrap_text=True)
        s3.merge_cells(start_row=r, start_column=col_start + 1, end_row=r, end_column=col_end)
        cl_lines = estimate_cl_rows_h(cl_text, cols_width_chars=(col_end - col_start) * 10)
        s3.row_dimensions[r].height = max(60, cl_lines * 14)
        pane_bg(s3, r, BG_ALT)
        r += 1

        # ── TMPL (frozen bid-time template from opp.originalStrategy) ──
        tmpl = card_data.get("cl_template_frozen") or card_data.get("cl_template_used")
        scn = card_data.get("scanner_name") or "—"
        tmpl_text = tmpl or "(no frozen template captured — manual bid, or opp predates the freeze)"
        tmpl_label_val = f"TMPL\n{len(tmpl) if tmpl else 0} ch\nscanner: {_clean(scn, 20)}"
        lc = s3.cell(row=r, column=col_start, value=tmpl_label_val)
        lc.font = fg_bold(OKAY_FG, 9); lc.alignment = Alignment(vertical="top", wrap_text=True)
        tc = s3.cell(row=r, column=col_start + 1, value=tmpl_text)
        tc.font = fg_mono(OKAY_FG if tmpl else FG_DIM, size=9)
        tc.alignment = Alignment(vertical="top", wrap_text=True)
        s3.merge_cells(start_row=r, start_column=col_start + 1, end_row=r, end_column=col_end)
        tmpl_lines = estimate_cl_rows_h(tmpl_text, cols_width_chars=(col_end - col_start) * 10)
        s3.row_dimensions[r].height = max(40, tmpl_lines * 14)
        pane_bg(s3, r, BG_ALT)
        return r

    pairs = (c2c.get("pairs") or [])
    for win_idx, pair in enumerate(pairs, 1):
        cw = pair["competitor_win"]
        uw = pair.get("ub_win_match") or None
        ul = pair.get("ub_loss_match") or None

        # --- win subheader (full width) ---
        sh_text = (f"  ▸ Win #{win_idx}  —  {_clean(cw.get('job_title'), 140) or '—'}  "
                   f"·  bid {_fmt_money(cw.get('bid_amount'))}  "
                   f"·  client {_fmt_money((cw.get('client_total_spent') or 0))}  "
                   f"·  {cw.get('connects_bid') or '—'} connects  ·  tier={cw.get('selection_tier')}")
        c = s3.cell(row=row, column=1, value=sh_text)
        c.font = fg_bold(FG, 11)
        c.alignment = Alignment(vertical="center")
        for cc in range(1, 15):
            s3.cell(row=row, column=cc).fill = fill(BG_ALT)
        s3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
        s3.row_dimensions[row].height = 22
        row += 1

        # --- Competitor win card + UB WIN card side-by-side (rows = 6) ---
        cw_tmpl, cw_scanner, cw_match_pct = _lookup_frozen_template_for_comp(cw.get("proposal_id"))

        cw_card_data = dict(cw)
        cw_card_data["cl_template_frozen"] = cw_tmpl
        cw_card_data["scanner_name"] = cw_scanner

        pair_block_start = row
        left_end = render_bid_card(row, 1, 6, "THEIR WIN", LINK_FG, cw_card_data, cl_color=FG)
        uw_badge_txt, uw_badge_fg = _match_badge(uw)
        if uw:
            right_end = render_bid_card(row, 8, 12, "UB WIN", WIN_FG, uw, cl_color=WIN_FG,
                                        match_badge_text=uw_badge_txt, match_badge_color=uw_badge_fg)
        else:
            # write a single stub row on the UB WIN side; won't align with left card
            # bottom but left card is 6 rows tall so we pad the right side to match.
            render_bid_card(row, 8, 12, "UB WIN", WIN_FG, {}, cl_color=FG_DIM,
                            is_missing_reason="— no comparable UB WIN above similarity threshold —")
            right_end = row  # single row
        # Align pane heights: pad the shorter side with empty rows so the block
        # ends cleanly at max(left_end, right_end).
        pair_ub_win_end = max(left_end, right_end)
        for rr in range(min(left_end, right_end) + 1, pair_ub_win_end + 1):
            pane_bg(s3, rr, BG_ALT)

        # --- AI REASONING in col N — merged across the full pair (spans UB WIN + UB LOSS) ---
        ai_www = pair.get("ai_what_worked_for_them") or ""
        ai_uwd = pair.get("ai_what_ubiquify_did") or ""
        ai_tactic = pair.get("ai_specific_tactic_to_copy") or ""
        rsn_parts = []
        if ai_www: rsn_parts.append("WHAT WORKED FOR THEM\n" + ai_www)
        if ai_uwd: rsn_parts.append("WHAT UBIQUIFY DID\n" + ai_uwd)
        if ai_tactic: rsn_parts.append("TACTIC TO COPY\n" + ai_tactic)
        rsn_full = "\n\n".join(rsn_parts) if rsn_parts else (pair.get("reasoning") or "—")

        reason_cell = s3.cell(row=pair_block_start, column=REASON_COL, value=rsn_full)
        reason_cell.font = fg(FG)
        reason_cell.alignment = Alignment(wrap_text=True, vertical="top")

        row = pair_ub_win_end + 1

        # --- UB LOSS card (right pane only, faded left) ---
        if ul:
            # small sep row
            for cc in range(1, 15):
                s3.cell(row=row, column=cc).fill = fill(BG)
            s3.row_dimensions[row].height = 4
            row += 1
            # faded "compare to their same win vs UB's losing bid" label on left
            s3.cell(row=row, column=1, value="↘ Same competitor win →").font = fg_bold(FG_DIM, 9)
            s3.cell(row=row, column=1).alignment = Alignment(vertical="top", wrap_text=True)
            s3.cell(row=row, column=2,
                    value=f"UB's CLOSEST LOSING bid on a similar job").font = fg(FG_DIM)
            s3.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            pane_bg(s3, row, BG_ALT)
            s3.row_dimensions[row].height = 16
            ul_badge_txt, ul_badge_fg = _match_badge(ul)
            # Start the loss card on the SAME row as the "↘ Same competitor win" label
            # on the right pane.
            # write UB LOSS card (JOB/CLIENT/BID/FL/CL/TMPL) from this row across cols 8-12
            ul_end = render_bid_card(row, 8, 12, "UB LOSS", CRIT_FG, ul, cl_color=CRIT_FG,
                                     match_badge_text=ul_badge_txt, match_badge_color=ul_badge_fg)
            # Fill the left side cells below the label row with faded BG_ALT + dim text
            for rr in range(row + 1, ul_end + 1):
                for cc in range(1, 7):
                    if s3.cell(row=rr, column=cc).fill.fill_type != "solid":
                        s3.cell(row=rr, column=cc).fill = fill(BG_ALT)
                    s3.cell(row=rr, column=cc).alignment = Alignment(vertical="top", wrap_text=True)
            row = ul_end + 1

        # Merge col N across the entire pair (from pair_block_start to row-1).
        pair_n_end_row = row - 1
        if pair_n_end_row > pair_block_start:
            s3.merge_cells(start_row=pair_block_start, start_column=REASON_COL,
                           end_row=pair_n_end_row, end_column=REASON_COL)

        # --- separator between wins ---
        s3.row_dimensions[row].height = 8
        for cc in range(1, 15):
            s3.cell(row=row, column=cc).fill = fill(BG)
        row += 1

    # ---- competitor separator ----
    for cc in range(1, 15):
        s3.cell(row=row, column=cc).fill = fill(BG)
    s3.row_dimensions[row].height = 14
    row += 1

# Fill rest of sheet bg + col widths
fill_sheet_bg(s3, max_row=row + 3, max_col=14)
# col widths: A=10 B=26 C=50 D=12 E=12 F=14 G=2 H=10 I=24 J=40 K=12 L=14 M=2 N=36
set_col_widths(s3, [18, 26, 50, 12, 12, 14, 2, 18, 24, 40, 12, 14, 2, 36])


# =========================================================================
# Sheet 4 — Chat Excerpts (skipped note)
# =========================================================================
s4 = wb.create_sheet("Chat Excerpts")
darken_sheet(s4)
s4["A1"] = "Chat Excerpts — SKIPPED"
s4["A1"].font = fg_bold(CRIT_FG, 16)
s4["A2"] = ("leads.chats collection is not populated for Ubiquify. 759 proposals have chat.chatId, "
            "but 0 leads.chats docs exist for this team. Chat-sync has not run.")
s4["A2"].font = fg(FG)
s4["A2"].alignment = Alignment(wrap_text=True)
s4["A3"] = ("Impact: we cannot diagnose the CL→reply→interview conversion step. Recommend "
            "running the chat-sync backfill for team 679a215568faa05722aabb93 and re-auditing "
            "in 2 weeks.")
s4["A3"].font = fg(FG_DIM)
s4["A3"].alignment = Alignment(wrap_text=True)
s4.row_dimensions[2].height = 32
s4.row_dimensions[3].height = 48
fill_sheet_bg(s4, max_row=20, max_col=5)
set_col_widths(s4, [100, 20, 20, 20, 20])


# =========================================================================
# Sheet 5 — Win/Loss CL Table
# =========================================================================
s5 = wb.create_sheet("Win-Loss CL Table")
darken_sheet(s5)
s5["A1"] = "Win / Loss CL Table — Paired per scanner (focus window)"
s5["A1"].font = fg_bold(FG, 16)
s5["A2"] = "Juxtapose winners against losers for the SAME scanner. The visual flip between rows is the analysis."
s5["A2"].font = fg(FG_DIM)

row = 4
wl_headers = [
    "Outcome", "Scanner", "Algorithm", "Prompt v", "Match %",
    "Job Title", "Hire Date / Created", "Client", "Client Spent",
    "Client FB Score", "Terms", "JD Excerpt (300ch)", "Rendered CL Excerpt (400ch)",
]
for i, h in enumerate(wl_headers, 1):
    c = s5.cell(row=row, column=i, value=h)
    c.font = fg_bold(FG_DIM)
    c.fill = fill(HEADER_BG)
row += 1

for sc in phase4["scanners"]:
    # Scanner header row
    s5.cell(row=row, column=1, value=f"━━ {sc['scanner_name']} ━━").font = fg_bold(FG, 12)
    s5.cell(row=row, column=2, value=f"sent in focus: {sc['sent_in_focus']}").font = fg(FG_DIM)
    cfg = sc.get("config") or {}
    q = (cfg.get("query_q") or "")[:200]
    s5.cell(row=row, column=3, value=f"query: {q}").font = fg_mono(FG_DIM)
    s5.merge_cells(start_row=row, start_column=3, end_row=row, end_column=13)
    fill_row_bg(s5, row, 1, 13, HEADER_BG)
    s5.row_dimensions[row].height = 24
    row += 1

    for r in sc["rows"]:
        outcome = r["outcome"]
        bg = WIN_BG if outcome == "WIN" else CRIT_BG
        fg_o = WIN_FG if outcome == "WIN" else CRIT_FG
        s5.cell(row=row, column=1, value=outcome).font = fg_bold(fg_o)
        s5.cell(row=row, column=2, value=r.get("scanner_name")).font = fg(FG)
        s5.cell(row=row, column=3, value=r.get("algorithm_signature") or "—").font = fg_mono(FG)
        s5.cell(row=row, column=4, value=r.get("prompt_version") or "—").font = fg_mono(FG)
        s5.cell(row=row, column=5, value=f"{r.get('match_percentage')}%" if r.get('match_percentage') else "—").font = fg(FG)
        s5.cell(row=row, column=6, value=(r.get("job_title") or "")[:80]).font = fg(FG)
        s5.cell(row=row, column=6).alignment = Alignment(wrap_text=True, vertical="top")
        dt = (r.get("hire_date") or r.get("meta_createdAt") or "")[:10]
        s5.cell(row=row, column=7, value=dt).font = fg(FG)
        s5.cell(row=row, column=8, value=(r.get("client_company") or "—")[:28]).font = fg(FG)
        cs = r.get("client_total_spent")
        s5.cell(row=row, column=9, value=f"${cs:,.0f}" if cs else "—").font = fg(FG)
        fb = r.get("client_feedback_score")
        s5.cell(row=row, column=10, value=f"{fb:.2f}" if fb else "—").font = fg(FG)
        hr = r.get("hourly_rate")
        am = r.get("amount")
        cb = r.get("connects_bid") or r.get("connects_expended")
        terms = f"{hr or am or ''}\nc={cb}"
        s5.cell(row=row, column=11, value=terms).font = fg_mono(FG)
        c = s5.cell(row=row, column=12, value=(r.get("jd_excerpt") or "")[:300])
        c.font = fg_mono(FG, size=8)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c = s5.cell(row=row, column=13, value=(r.get("rendered_cover_letter_excerpt") or "")[:400])
        c.font = fg_mono(FG, size=8)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        # band outcome cell
        s5.cell(row=row, column=1).fill = fill(bg)
        for cc in range(2, 14):
            s5.cell(row=row, column=cc).fill = fill(BG_ALT)
        s5.row_dimensions[row].height = 140
        row += 1

fill_sheet_bg(s5, max_row=row+3, max_col=13)
set_col_widths(s5, [10, 22, 14, 14, 8, 32, 14, 22, 12, 10, 14, 50, 60])


# =========================================================================
# Sheet 6 — Auto-Bidding Aggregates
# =========================================================================
s6 = wb.create_sheet("Auto-Bidding Aggregates")
darken_sheet(s6)
s6["A1"] = "Auto-Bidding Aggregates — scanner × algorithm × prompt version"
s6["A1"].font = fg_bold(FG, 16)

row = 3
agg_headers = [
    "Scanner", "Algorithm", "Prompt v", "Sent", "Replies", "Reply Rate",
    "Hires (diag)", "Connects", "$/Reply", "Prior RR", "Δ RR", "Status",
]
for i, h in enumerate(agg_headers, 1):
    c = s6.cell(row=row, column=i, value=h)
    c.font = fg_bold(FG_DIM)
    c.fill = fill(HEADER_BG)
row += 1

# Show slices with sent ≥ 5 in focus
team_avg_rr = phase1["focus_window"]["auto_bidder"]["reply_rate"]
p75_rr = phase2a["inferred_cohort"]["reply_rate_stats"]["p75"]
for r in phase5["focus_rows"]:
    if r["sent"] < 5:
        continue
    s6.cell(row=row, column=1, value=r.get("scanner_name") or r["scanner_id"][:8]).font = fg(FG)
    s6.cell(row=row, column=2, value=r.get("algorithm_signature") or "—").font = fg_mono(FG)
    s6.cell(row=row, column=3, value=r.get("prompt_version") or "—").font = fg_mono(FG)
    s6.cell(row=row, column=4, value=r["sent"]).font = fg(FG)
    s6.cell(row=row, column=5, value=r["replies"]).font = fg(FG)
    rr = r.get("reply_rate") or 0
    s6.cell(row=row, column=6, value=f"{rr*100:.2f}%").font = fg(FG)
    s6.cell(row=row, column=7, value=r["hires"]).font = fg(FG)
    s6.cell(row=row, column=8, value=r["connects"]).font = fg(FG)
    cpr = r.get("cost_per_reply_usd")
    s6.cell(row=row, column=9, value=f"${cpr:.2f}" if cpr else "—").font = fg(FG)
    pr = r.get("prior_reply_rate")
    s6.cell(row=row, column=10, value=f"{pr*100:.2f}%" if pr is not None else "—").font = fg(FG)
    d = r.get("delta_reply_rate")
    s6.cell(row=row, column=11, value=f"{d*100:+.2f}pp" if d is not None else "—").font = fg(FG)
    # Status band
    if rr >= p75_rr and r["sent"] >= 30:
        status, bg = "GREEN (>p75)", WIN_BG
    elif rr == 0 and r["sent"] >= 15:
        status, bg = "RED (0 RR, wasted)", CRIT_BG
    elif rr < team_avg_rr * 0.6 and r["sent"] >= 15:
        status, bg = "AMBER (below avg×0.6)", OKAY_BG
    else:
        status, bg = "", BG_ALT
    c = s6.cell(row=row, column=12, value=status)
    c.font = fg_bold(FG, 9)
    if status:
        c.fill = fill(bg)
        for cc in range(1, 12):
            s6.cell(row=row, column=cc).fill = fill(BG_ALT)
    else:
        for cc in range(1, 13):
            s6.cell(row=row, column=cc).fill = fill(BG_ALT)
    row += 1

fill_sheet_bg(s6, max_row=row+5, max_col=12)
set_col_widths(s6, [26, 16, 14, 8, 8, 10, 10, 10, 10, 10, 10, 20])


# =========================================================================
# Sheet 7 — Recommendations Detail
# =========================================================================
s7 = wb.create_sheet("Recommendations Detail")
darken_sheet(s7)
s7["A1"] = "Recommendations Detail — every Tier 2 & Tier 3 item, operationalized"
s7["A1"].font = fg_bold(FG, 16)

row = 3
rec_headers = ["Tier", "Lever", "Specific Change", "Evidence", "Expected Effect",
               "Success Metric", "Rollout Scope", "Duration", "Min Sample", "Stop-Loss"]
for i, h in enumerate(rec_headers, 1):
    c = s7.cell(row=row, column=i, value=h)
    c.font = fg_bold(FG_DIM)
    c.fill = fill(HEADER_BG)
row += 1

recs = [
    ("CRITICAL", "Pause scanner",
     "Pause D3 Awais - Full Stack & AI scanner. Rewrite boolean query or merge into A4.",
     "Phase 5 row: D3 Awais 21 sent / 0% RR, 2 consecutive windows at 0%.",
     "Recover ~$63/mo of connect spend; redirect to a scanner with ≥4% RR.",
     "$/reply below team avg AFTER rewrite",
     "D3 Awais only", "immediate", "20 proposals post-rewrite",
     "If rewritten scanner still <2% after 20 sent, retire permanently."),
    ("CRITICAL", "Revert scanner config",
     "Roll back AM-07-D to the -C sibling baseline config until the -D variant can be re-tested.",
     "Phase 5: AM-07-D 15 sent 0% RR vs prior 2.33%.",
     "Restore 2%+ reply rate on this scanner.",
     "Reply rate matches prior window",
     "AM-07 only", "30 days", "30 proposals",
     "If reverted config doesn't recover RR, this scanner's target space has changed."),
    ("CRITICAL", "Revert scanner config",
     "Roll back AM-05-A and AM-03-C to their previous-window configs.",
     "Phase 5: both dropped from 12% and 8.33% to 0% RR.",
     "Recover the lost reply rate.",
     "RR matches prior window",
     "AM-05-A, AM-03-C", "30 days", "25 proposals each",
     "If still 0 after 25 sent, suspect broader platform or targeting shift."),
    ("CRITICAL", "Data hygiene",
     "Populate teams.serviceNames = ['Web/Mobile/SW Dev', 'AI/ML'] and teams.industry for Ubiquify.",
     "Phase 1: team doc has serviceNames=[] and industry=null → breaks internal benchmark widgets.",
     "Enables dashboard.benchmarks lookup; correct peer cohort for all future audits.",
     "Benchmark widgets populate",
     "team doc only", "one-time", "n/a",
     "n/a"),
    ("CRITICAL", "Platform engineering",
     "Run chat-sync backfill for team 679a215568faa05722aabb93.",
     "Phase 3: 759 proposals with chatId but 0 leads.chats — CL→interview funnel unmeasurable.",
     "Enables Section 3 diagnostics on next audit; unlocks chat-driven retention features.",
     "leads.chats populated for team",
     "GigRadar infra", "1 day", "n/a",
     "n/a"),
    ("OKAY", "Connects cap",
     "Cap connectsBid at 12 on scanners with <4% reply rate in the last 60 days (V. python, A4 Awais, AM-07).",
     "Phase 5: these scanners spend $60–128 per reply — 2-3x peer median $36.",
     "Cut $/reply ~30% on those scanners without losing much volume.",
     "$/reply below team avg",
     "3 scanners", "45 days", "30 proposals per scanner",
     "If reply rate drops >1pp, restore prior bid cap."),
    ("OKAY", "Algo/prompt rollout",
     "A/B test V. python + A4 Awais with the ALG_LAZ + promptVersion=1.2.6.2.mem combo (same as top-performing AM-05-C).",
     "Phase 5: AM-05-C with ALG_LAZ+1.2.6.2.mem hits 10.20% RR; V. python and A4 Awais still on older 'ㅤ⁤'+1.2.6.1.",
     "Match AM-05-C's 10% RR on adjacent scanners.",
     "Delta RR vs prior window ≥ +3pp",
     "50% of V. python + A4 Awais volume", "30 days", "60 proposals per scanner-arm",
     "If new config is ≥1pp WORSE than old, roll back."),
    ("OKAY", "CL template — opening hook",
     "Add a 'quick call' CTA with Calendly link to the first chat-bot auto-reply template.",
     "Phase 2B competitors (RipeSeed, WeSoftYou, Vizio) include call-booking link in first reply; converts chats to interviews.",
     "Lift hire rate from 0.59% toward peer p75 (1.3%).",
     "Interviewed-on-reply rate",
     "all scanners", "30 days", "20 replies handled",
     "n/a"),
    ("OKAY", "Account positioning",
     "Populate Ubiquify's Upwork agency profile title + description. Currently empty.",
     "Phase 1: agency profile has no title/description/category/skills populated. Competitors (RipeSeed, League Design, Perfsol, WeSoftYou) all have rich positioning profiles.",
     "Higher client click-through from proposal to profile; more invites.",
     "Invite count delta +20%",
     "agency profile", "one-time", "measure at 45d",
     "n/a"),
]
for r in recs:
    tier = r[0]
    bg = WIN_BG if tier == "WIN" else (CRIT_BG if tier == "CRITICAL" else OKAY_BG)
    fg_c = WIN_FG if tier == "WIN" else (CRIT_FG if tier == "CRITICAL" else OKAY_FG)
    s7.cell(row=row, column=1, value=tier).font = fg_bold(fg_c)
    s7.cell(row=row, column=1).fill = fill(bg)
    for i, v in enumerate(r[1:], 2):
        c = s7.cell(row=row, column=i, value=v)
        c.font = fg(FG)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.fill = fill(BG_ALT)
    s7.row_dimensions[row].height = 84
    row += 1

fill_sheet_bg(s7, max_row=row+3, max_col=10)
set_col_widths(s7, [10, 18, 48, 42, 32, 18, 20, 12, 16, 26])


# ---- post-process: force col A to vertical=top across all sheets ----
# Also ensure any cell in any column that has no explicit alignment gets vertical=top.
for ws in wb.worksheets:
    max_r = ws.max_row
    max_c = ws.max_column
    for r in range(1, max_r + 1):
        # Col A specifically: always vertical=top, preserve wrap_text if already set
        cell = ws.cell(row=r, column=1)
        existing_wrap = cell.alignment.wrap_text if cell.alignment else True
        existing_horiz = cell.alignment.horizontal if cell.alignment else None
        cell.alignment = Alignment(
            wrap_text=existing_wrap if existing_wrap is not None else True,
            vertical="top",
            horizontal=existing_horiz,
        )

# ---- save ----
OUT = "/sessions/dazzling-nifty-fermat/mnt/GigRadar AI Auto Researcher/Ubiquify_Audit_2026-04-22.xlsx"
wb.save(OUT)
print(f"Saved: {OUT}")
