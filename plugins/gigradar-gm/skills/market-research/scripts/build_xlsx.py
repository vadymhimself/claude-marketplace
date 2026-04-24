"""Build a GigRadar market-research Excel workbook from tidy CSVs.

Inputs (produced by merge_all.py):
  <tidy>/category__full.csv
  <tidy>/subcategory__full.csv
  <tidy>/skill__full.csv

Sheets:
  - Overview          (market totals + top 5 categories + market-wide reply rate)
  - Categories        (full trio: volume / quality / reply for each category)
  - Subcategories
  - Top skills        (filtered min-volume, top 150)
  - Trending up       (biggest latest-vs-prior gains, min floor)
  - Trending down     (biggest losses)
  - Reply leaders     (cross-dim, highest reply rate, min 100 proposals)
  - Methodology

The "focus window" is the LAST window in --windows; the "prior window" is the
second-to-last. Deltas are computed latest-vs-prior. Overview totals and reply
rates are filled in for every window passed.

Usage:
  python build_xlsx.py --tidy tidy/ --out path/to/report.xlsx \
                       --windows apr2025,may2025,jun2025 --focus may2025 \
                       [--market-totals '{"apr2025":198783,"may2025":197724}'] \
                       [--market-reply '{"may2025":[93487,9934,28081]}']
"""
from __future__ import annotations
import argparse, csv, json, os, sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not installed. `pip install openpyxl --break-system-packages`", file=sys.stderr)
    raise

HEADER_FILL = PatternFill("solid", start_color="1F3A5F")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
BODY_FONT   = Font(name="Arial", size=10)
TITLE_FONT  = Font(bold=True, size=14, name="Arial", color="1F3A5F")
SUB_FONT    = Font(italic=True, size=10, name="Arial", color="555555")
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def pretty_window(w: str) -> str:
    """'may2025' -> 'May 2025'. Leaves other forms untouched."""
    if len(w) == 7 and w[:3].isalpha() and w[3:].isdigit():
        return f"{w[:3].title()} {w[3:]}"
    return w


def load(tidy_dir: str, dim: str) -> list[dict]:
    path = os.path.join(tidy_dir, f"{dim}__full.csv")
    if not os.path.exists(path):
        print(f"  missing: {path}", file=sys.stderr)
        return []
    with open(path) as f:
        rows = list(csv.DictReader(f))
    for r in rows:
        for k, v in list(r.items()):
            if k == "key" or v == "" or v is None:
                continue
            try:
                r[k] = float(v)
                if r[k].is_integer():
                    r[k] = int(r[k])
            except (ValueError, TypeError):
                pass
    return rows


def safe_float(v, default=None):
    try: return float(v)
    except (ValueError, TypeError): return default


def style_header(ws, row, cols):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def write_header(ws, row, headers):
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header(ws, row, range(1, len(headers) + 1))
    ws.row_dimensions[row].height = 32


def build_main_sheet(ws, rows: list[dict], name: str, focus: str, prior: str | None):
    ws.title = name
    ws["A1"] = f"Upwork — {name} — {pretty_window(focus)} focus"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=18)
    ws["A2"] = ("Volume & quality from ES metajob (public crawl). Reply rate = proposals.dashroomUID non-null"
                " (canonical formula, StatsRepository). Base: meta.createdAt in window, inviteToInterviewUid = null.")
    ws["A2"].font = SUB_FONT
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=18)

    fw, pw = focus, prior
    delta_jobs_col = f"Jobs Δ {pretty_window(fw)}/{pretty_window(pw)}" if pw else f"Jobs Δ (vs prior)"
    delta_reply_col = f"Δ reply rate {pretty_window(fw)}/{pretty_window(pw)} (abs pp)" if pw else "Δ reply rate (abs pp)"

    headers = [
        "Key",
        f"Jobs ({pretty_window(fw)})",
        f"Jobs ({pretty_window(pw)})" if pw else "Jobs (prior)",
        delta_jobs_col,
        "% Hourly", "Median $/hr (min)", "Median $/hr (max)",
        "Median $ fixed", "P25 fixed", "P75 fixed",
        "Avg client total spent ($)", "% Payment verified",
        "Avg hire rate", "Avg feedback",
        f"Proposals ({pretty_window(fw)})", "Reply rate", "View rate",
        delta_reply_col,
    ]
    start_row = 4
    write_header(ws, start_row, headers)

    for i, r in enumerate(rows, start=start_row + 1):
        ws.cell(row=i, column=1, value=r.get("key"))
        delta_jobs_key = f"{fw}_vs_{pw}_jobs_pct" if pw else None
        delta_reply_key = f"{fw}_vs_{pw}_reply_rate_abs" if pw else None
        vals = [
            r.get(f"{fw}_jobs"),
            r.get(f"{pw}_jobs") if pw else None,
            r.get(delta_jobs_key) if delta_jobs_key else None,
            r.get(f"{fw}_pct_hourly"),
            r.get(f"{fw}_median_hourly_min"),
            r.get(f"{fw}_median_hourly_max"),
            r.get(f"{fw}_median_fixed"),
            r.get(f"{fw}_p25_fixed"),
            r.get(f"{fw}_p75_fixed"),
            r.get(f"{fw}_avg_total_spent"),
            r.get(f"{fw}_pct_pv"),
            r.get(f"{fw}_avg_hire_rate"),
            r.get(f"{fw}_avg_feedback"),
            r.get(f"{fw}_proposals"),
            r.get(f"{fw}_reply_rate"),
            r.get(f"{fw}_view_rate"),
            r.get(delta_reply_key) if delta_reply_key else None,
        ]
        for j, v in enumerate(vals, start=2):
            cell = ws.cell(row=i, column=j, value=(None if v in ("", None) else v))
            cell.font = BODY_FONT

    widths = [36, 12, 12, 14, 10, 14, 14, 14, 12, 12, 22, 16, 13, 13, 14, 13, 13, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    last_row = start_row + len(rows)
    for r in range(start_row + 1, last_row + 1):
        for c in [2, 3]:   # jobs counts
            ws.cell(row=r, column=c).number_format = '#,##0'
        for c in [4]:      # jobs delta pct
            ws.cell(row=r, column=c).number_format = '+0.0%;-0.0%;-'
        for c in [5, 12, 16, 17]:  # percentages
            ws.cell(row=r, column=c).number_format = '0.0%;-;-'
        for c in [6, 7, 8, 9, 10]:  # dollar amounts
            ws.cell(row=r, column=c).number_format = '$#,##0.00;($#,##0.00);-'
        ws.cell(row=r, column=11).number_format = '$#,##0;($#,##0);-'
        ws.cell(row=r, column=13).number_format = '0.0%;-;-'
        ws.cell(row=r, column=14).number_format = '0.0;-;-'
        ws.cell(row=r, column=15).number_format = '#,##0'
        ws.cell(row=r, column=18).number_format = '+0.0%;-0.0%;-'

    if last_row >= start_row + 1:
        ws.conditional_formatting.add(
            f"R{start_row+1}:R{last_row}",
            ColorScaleRule(start_type='min', start_color='FFFFFF', end_type='max', end_color='63BE7B')
        )
        ws.conditional_formatting.add(
            f"D{start_row+1}:D{last_row}",
            ColorScaleRule(start_type='min', start_color='F8696B',
                           mid_type='num', mid_value=0, mid_color='FFEB84',
                           end_type='max', end_color='63BE7B')
        )
    ws.freeze_panes = f"B{start_row+1}"


def build_trending(ws, rows, title: str, direction: str, focus: str, prior: str,
                   floor_jobs: int = 500, limit: int = 25):
    ws.title = f"Trending {direction}"
    ws["A1"] = f"{title} — {pretty_window(focus)} vs {pretty_window(prior)} (min {floor_jobs:,} jobs in {pretty_window(focus)})"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=7)

    headers = ["Key", f"Jobs ({pretty_window(prior)})", f"Jobs ({pretty_window(focus)})",
               f"Δ {pretty_window(focus)}/{pretty_window(prior)}",
               f"Reply rate ({pretty_window(focus)})", f"Avg client spent ($, {pretty_window(focus)})"]
    write_header(ws, 3, headers)

    delta_key = f"{focus}_vs_{prior}_jobs_pct"
    filtered = [r for r in rows
                if safe_float(r.get(f"{focus}_jobs"), 0) >= floor_jobs
                and safe_float(r.get(delta_key)) is not None]
    filtered.sort(key=lambda r: safe_float(r.get(delta_key), 0), reverse=(direction == "up"))

    for i, r in enumerate(filtered[:limit], start=4):
        ws.cell(row=i, column=1, value=r.get("key"))
        ws.cell(row=i, column=2, value=safe_float(r.get(f"{prior}_jobs")))
        ws.cell(row=i, column=3, value=safe_float(r.get(f"{focus}_jobs")))
        ws.cell(row=i, column=4, value=safe_float(r.get(delta_key)))
        ws.cell(row=i, column=5, value=safe_float(r.get(f"{focus}_reply_rate")))
        ws.cell(row=i, column=6, value=safe_float(r.get(f"{focus}_avg_total_spent")))
        for c in [2, 3]:
            ws.cell(row=i, column=c).number_format = '#,##0'
        ws.cell(row=i, column=4).number_format = '+0.0%;-0.0%;-'
        ws.cell(row=i, column=5).number_format = '0.0%;-;-'
        ws.cell(row=i, column=6).number_format = '$#,##0;-;-'
        for c in range(1, 7):
            ws.cell(row=i, column=c).font = BODY_FONT

    widths = [40, 14, 14, 14, 16, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "B4"


def build_reply_leaders(ws, all_rows_by_dim: dict, focus: str, prior: str | None, min_proposals: int = 100):
    ws.title = "Reply rate leaders"
    ws["A1"] = f"Highest reply rate — {pretty_window(focus)} (min {min_proposals} proposals, all dimensions)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=8)
    ws["A2"] = "Reply = proposal.dashroomUID non-null. Data from GigRadar customer proposals pool."
    ws["A2"].font = SUB_FONT
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=8)

    delta_col = f"Δ reply {pretty_window(focus)}/{pretty_window(prior)}" if prior else "Δ reply (vs prior)"
    headers = ["Dimension", "Key", f"Jobs ({pretty_window(focus)})", "Proposals", "Replies",
               "Reply rate", "View rate", delta_col]
    write_header(ws, 4, headers)

    delta_key = f"{focus}_vs_{prior}_reply_rate_abs" if prior else None
    unioned = []
    for dim, rows in all_rows_by_dim.items():
        for r in rows:
            p = safe_float(r.get(f"{focus}_proposals"), 0)
            if p < min_proposals:
                continue
            unioned.append({
                "dim": dim,
                "key": r["key"],
                "jobs": safe_float(r.get(f"{focus}_jobs"), 0),
                "proposals": p,
                "replies": safe_float(r.get(f"{focus}_replies"), 0),
                "reply_rate": safe_float(r.get(f"{focus}_reply_rate")),
                "view_rate":  safe_float(r.get(f"{focus}_view_rate")),
                "delta": safe_float(r.get(delta_key)) if delta_key else None,
            })
    unioned.sort(key=lambda r: r["reply_rate"] or 0, reverse=True)

    for i, r in enumerate(unioned[:40], start=5):
        ws.cell(row=i, column=1, value=r["dim"])
        ws.cell(row=i, column=2, value=r["key"])
        ws.cell(row=i, column=3, value=r["jobs"])
        ws.cell(row=i, column=4, value=r["proposals"])
        ws.cell(row=i, column=5, value=r["replies"])
        ws.cell(row=i, column=6, value=r["reply_rate"])
        ws.cell(row=i, column=7, value=r["view_rate"])
        ws.cell(row=i, column=8, value=r["delta"])
        ws.cell(row=i, column=3).number_format = '#,##0'
        ws.cell(row=i, column=4).number_format = '#,##0'
        ws.cell(row=i, column=5).number_format = '#,##0'
        ws.cell(row=i, column=6).number_format = '0.0%'
        ws.cell(row=i, column=7).number_format = '0.0%'
        ws.cell(row=i, column=8).number_format = '+0.0%;-0.0%;-'
        for c in range(1, 9):
            ws.cell(row=i, column=c).font = BODY_FONT

    widths = [13, 42, 14, 12, 10, 12, 12, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C5"


def build_overview(ws, cat_rows, windows, focus, market_totals: dict, market_reply: dict):
    ws.title = "Overview"
    title_bits = "/".join(pretty_window(w) for w in windows)
    ws["A1"] = f"Upwork market — {title_bits}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=6)

    ws["A3"] = "Market totals (from ES metajob public crawl)"
    ws["A3"].font = Font(bold=True, name="Arial", size=12)
    write_header(ws, 4, ["Window", "Jobs posted", "Δ vs prior"])
    prev_n = None
    for i, w in enumerate(windows, start=5):
        n = market_totals.get(w)
        ws.cell(row=i, column=1, value=pretty_window(w))
        ws.cell(row=i, column=2, value=n if n is not None else None)
        ws.cell(row=i, column=2).number_format = '#,##0'
        if prev_n and n:
            ws.cell(row=i, column=3, value=(n - prev_n) / prev_n)
            ws.cell(row=i, column=3).number_format = '+0.0%;-0.0%;-'
        for c in range(1, 4):
            ws.cell(row=i, column=c).font = BODY_FONT
        prev_n = n

    start = 6 + len(windows)
    ws.cell(row=start, column=1, value="Market-wide reply / view rates (from GigRadar proposals pool)")
    ws.cell(row=start, column=1).font = Font(bold=True, name="Arial", size=12)
    write_header(ws, start + 1, ["Window", "Proposals submitted", "Replies", "Reply rate", "Views", "View rate"])
    for i, w in enumerate(windows, start=start + 2):
        triple = market_reply.get(w) or [None, None, None]
        p, r, v = triple
        ws.cell(row=i, column=1, value=pretty_window(w))
        ws.cell(row=i, column=2, value=p)
        ws.cell(row=i, column=3, value=r)
        if p:
            ws.cell(row=i, column=4, value=f"=C{i}/B{i}")
        ws.cell(row=i, column=5, value=v)
        if p:
            ws.cell(row=i, column=6, value=f"=E{i}/B{i}")
        for c in [2, 3, 5]:
            ws.cell(row=i, column=c).number_format = '#,##0'
        for c in [4, 6]:
            ws.cell(row=i, column=c).number_format = '0.0%'
        for c in range(1, 7):
            ws.cell(row=i, column=c).font = BODY_FONT

    top_start = start + 2 + len(windows) + 1
    ws.cell(row=top_start, column=1, value=f"Top 5 categories ({pretty_window(focus)} share of jobs)")
    ws.cell(row=top_start, column=1).font = Font(bold=True, name="Arial", size=12)
    write_header(ws, top_start + 1, ["Category", f"Jobs ({pretty_window(focus)})", "Share", "Reply rate", "Avg client total spent"])
    cat_sorted = sorted(cat_rows, key=lambda r: safe_float(r.get(f"{focus}_jobs"), 0), reverse=True)[:5]
    for i, r in enumerate(cat_sorted, start=top_start + 2):
        ws.cell(row=i, column=1, value=r["key"])
        ws.cell(row=i, column=2, value=safe_float(r.get(f"{focus}_jobs")))
        ws.cell(row=i, column=3, value=safe_float(r.get(f"{focus}_jobs_share")))
        ws.cell(row=i, column=4, value=safe_float(r.get(f"{focus}_reply_rate")))
        ws.cell(row=i, column=5, value=safe_float(r.get(f"{focus}_avg_total_spent")))
        ws.cell(row=i, column=2).number_format = '#,##0'
        ws.cell(row=i, column=3).number_format = '0.0%'
        ws.cell(row=i, column=4).number_format = '0.0%'
        ws.cell(row=i, column=5).number_format = '$#,##0'
        for c in range(1, 6):
            ws.cell(row=i, column=c).font = BODY_FONT

    widths = [38, 22, 20, 14, 22, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_methodology(ws, windows, focus):
    ws.title = "Methodology"
    ws["A1"] = "Methodology & caveats"
    ws["A1"].font = TITLE_FONT

    windows_str = ", ".join(pretty_window(w) for w in windows)
    notes = [
        ("Data sources",
         "Volume & quality: Elasticsearch `metajob` index (public Upwork crawl)."
         " Reply/view rates: Mongo `proposals` collection aggregated across all GigRadar customer teams."),
        ("Windows",
         f"Reported windows: {windows_str}. Focus window: {pretty_window(focus)}. All bounds are [start, end) UTC."),
        ("Quality metrics",
         "Budget medians are computed separately for hourly (`budget.type=2`) and fixed (`budget.type=1`) jobs — mixing them is meaningless."
         " Client stats (`totalSpent`, `feedbackScore`, `hireRate`, `paymentVerified`) are averages over all jobs in the bucket."),
        ("Reply rate formula (canonical)",
         "Reply = `proposal.dashroomUID` non-null. Transcribed from `StatsRepository.getOpportunityStats`."
         " View = dashroomUID OR status==7 OR otherAnnotations contains 12."
         " Base filter: `meta.createdAt` in window AND `meta.inviteToInterviewUid` is null."),
        ("Sample-size biases",
         "Proposals are agency-side — there is one row per GigRadar customer who submitted. Categories under-represented"
         " among GigRadar customers (Legal, Translation, Customer Service) will have thin per-category samples."
         " Rows with <100 proposals have wide confidence intervals — treat as directional."),
        ("Category=None rows",
         "A subset of proposals have `metaJob.categoryName=null` (older records pre-dating the embedded metaJob pipeline)."
         " They are omitted from per-category reply-rate tables but included in market totals."),
        ("Indexing / perf",
         "Cross-team reply-rate aggregation requires explicit `hint={\"meta.createdAt\": 1}` — without it the planner picks"
         " the compound tenant index which is inefficient for date-first scans. Logged in DATA_REFERENCE.md §14."),
    ]
    for i, (h, body) in enumerate(notes, start=3):
        ws.cell(row=i*2, column=1, value=h).font = Font(bold=True, name="Arial", size=11, color="1F3A5F")
        c = ws.cell(row=i*2 + 1, column=1, value=body)
        c.font = BODY_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=i*2 + 1, end_row=i*2 + 1, start_column=1, end_column=6)
        ws.row_dimensions[i*2 + 1].height = 48

    ws.column_dimensions["A"].width = 32
    for c in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[c].width = 22


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--tidy", required=True, help="dir with <dim>__full.csv files")
    ap.add_argument("--out", required=True, help="output .xlsx path")
    ap.add_argument("--windows", required=True, help="ordered, comma-separated window slugs (must match merge_all windows)")
    ap.add_argument("--focus", default=None, help="focus window slug; default = last in --windows")
    ap.add_argument("--market-totals", default=None,
                    help='optional JSON: {"may2025": 197724, ...} — ES market totals for Overview sheet')
    ap.add_argument("--market-reply", default=None,
                    help='optional JSON: {"may2025": [proposals, replies, views], ...} — market-wide reply/view totals')
    ap.add_argument("--skill-min-jobs", type=int, default=200,
                    help="min focus-window jobs for a skill to appear in Top skills sheet")
    ap.add_argument("--skill-limit", type=int, default=150)
    ap.add_argument("--trend-floor", type=int, default=500)
    args = ap.parse_args()

    windows = [w.strip() for w in args.windows.split(",") if w.strip()]
    if len(windows) < 1:
        raise SystemExit("need at least one window")
    focus = args.focus or windows[-1]
    if focus not in windows:
        raise SystemExit(f"focus {focus!r} not in windows {windows}")
    prior = windows[windows.index(focus) - 1] if windows.index(focus) > 0 else None

    market_totals = json.loads(args.market_totals) if args.market_totals else {}
    market_reply  = json.loads(args.market_reply)  if args.market_reply else {}

    cat_rows   = load(args.tidy, "category")
    sub_rows   = load(args.tidy, "subcategory")
    skill_rows = load(args.tidy, "skill")
    skill_rows = [r for r in skill_rows if safe_float(r.get(f"{focus}_jobs"), 0) >= args.skill_min_jobs][:args.skill_limit]

    wb = Workbook()
    build_overview(wb.active, cat_rows, windows, focus, market_totals, market_reply)
    build_main_sheet(wb.create_sheet(), cat_rows,   "Categories",   focus, prior)
    build_main_sheet(wb.create_sheet(), sub_rows,   "Subcategories", focus, prior)
    build_main_sheet(wb.create_sheet(), skill_rows, "Top skills",   focus, prior)
    if prior:
        build_trending(wb.create_sheet(), skill_rows, "Skills trending up",   "up",   focus, prior, args.trend_floor, 25)
        build_trending(wb.create_sheet(), skill_rows, "Skills trending down", "down", focus, prior, args.trend_floor, 25)
        build_trending(wb.create_sheet(), sub_rows,   "Subcategories trending up", "up", focus, prior, args.trend_floor, 20)
    build_reply_leaders(wb.create_sheet(),
                        {"category": cat_rows, "subcategory": sub_rows, "skill": skill_rows},
                        focus, prior)
    build_methodology(wb.create_sheet(), windows, focus)

    os.makedirs(os.path.dirname(args.out) or ".", exist_ok=True)
    wb.save(args.out)
    print(f"wrote -> {args.out}")


if __name__ == "__main__":
    main()
